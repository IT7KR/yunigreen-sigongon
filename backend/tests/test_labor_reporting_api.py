"""Labor reporting API tests for site/monthly/welfare/admin-rate workflows."""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlmodel import select

from app.core.database import get_async_db
from app.core.security import create_access_token
from app.core.security import get_current_user
from app.main import app
from app.models.consent import ConsentRecord
from app.models.operations import ActivityLog, DailyWorker, InsuranceRate, WorkRecord
from app.models.project import Project
from app.models.user import Organization, User, UserRole


@pytest.fixture
async def labor_api_context(test_db):
    org = Organization(name="유니그린개발")
    test_db.add(org)
    await test_db.flush()

    user = User(
        username="company_admin_test",
        name="테스트관리자",
        role=UserRole.COMPANY_ADMIN,
        organization_id=org.id,
        password_hash="test_hash",
        is_active=True,
    )
    test_db.add(user)
    await test_db.commit()

    async def _fake_user():
        return user

    async def _override_db():
        try:
            yield test_db
            await test_db.commit()
        except Exception:
            await test_db.rollback()
            raise

    app.dependency_overrides[get_current_user] = _fake_user
    app.dependency_overrides[get_async_db] = _override_db

    yield {"org": org, "user": user, "db": test_db}

    app.dependency_overrides.pop(get_current_user, None)
    app.dependency_overrides.pop(get_async_db, None)


async def _seed_project_and_worker(
    context,
    *,
    job_type_code: str = "706",
    has_id_card: bool = True,
    has_safety_cert: bool = True,
    invite_token: str | None = None,
):
    db = context["db"]
    org = context["org"]
    user = context["user"]

    project = Project(
        name="OO중학교 방수공사",
        address="서울시",
        organization_id=org.id,
        created_by=user.id,
        started_at=None,
        completed_at=None,
    )
    db.add(project)
    await db.flush()

    worker = DailyWorker(
        organization_id=org.id,
        name="홍길동",
        job_type="보통인부",
        job_type_code=job_type_code,
        team="A팀",
        hire_date=date(2026, 1, 1),
        birth_date="900101",
        gender=1,
        address="서울시",
        daily_rate=Decimal("200000"),
        phone="010-1234-5678",
        is_foreign=False,
        has_id_card=has_id_card,
        has_safety_cert=has_safety_cert,
        invite_token=invite_token,
    )
    db.add(worker)
    await db.flush()
    await db.commit()

    return project, worker


@pytest.mark.asyncio
async def test_consolidated_report_merges_days_across_projects(async_client: AsyncClient, labor_api_context):
    db = labor_api_context["db"]
    org = labor_api_context["org"]

    project_a = Project(name="A현장", address="서울", organization_id=org.id)
    project_b = Project(name="B현장", address="경기", organization_id=org.id)
    db.add(project_a)
    db.add(project_b)
    await db.flush()

    worker = DailyWorker(
        organization_id=org.id,
        name="통합근로자",
        job_type="보통인부",
        job_type_code="706",
        team="통합",
        hire_date=date(2026, 1, 1),
        birth_date="900101",
        gender=1,
        daily_rate=Decimal("100000"),
        phone="01011112222",
    )
    db.add(worker)
    await db.flush()

    for day in range(11, 16):
        db.add(
            WorkRecord(
                worker_id=worker.id,
                project_id=project_a.id,
                work_date=date(2026, 1, day),
                man_days=Decimal("1"),
            )
        )
    for day in range(16, 21):
        db.add(
            WorkRecord(
                worker_id=worker.id,
                project_id=project_b.id,
                work_date=date(2026, 1, day),
                man_days=Decimal("1"),
            )
        )
    await db.commit()

    response = await async_client.get("/api/v1/labor/reports/consolidated", params={"year": 2026, "month": 1})
    assert response.status_code == 200

    payload = response.json()["data"]
    assert len(payload["entries"]) == 1

    entry = payload["entries"][0]
    assert entry["worker_name"] == "통합근로자"
    assert entry["total_days"] == 10
    assert entry["total_man_days"] == 10.0

    work_days = {int(day): value for day, value in entry["work_days"].items()}
    assert set(work_days.keys()) == set(range(11, 21))


@pytest.mark.asyncio
async def test_site_daily_labor_download_splits_by_month(async_client: AsyncClient, labor_api_context):
    db = labor_api_context["db"]
    project, worker = await _seed_project_and_worker(labor_api_context)

    db.add(
        WorkRecord(
            worker_id=worker.id,
            project_id=project.id,
            work_date=date(2026, 1, 20),
            man_days=Decimal("1"),
        )
    )
    db.add(
        WorkRecord(
            worker_id=worker.id,
            project_id=project.id,
            work_date=date(2026, 1, 21),
            man_days=Decimal("1"),
        )
    )
    db.add(
        WorkRecord(
            worker_id=worker.id,
            project_id=project.id,
            work_date=date(2026, 2, 14),
            man_days=Decimal("1"),
        )
    )
    await db.commit()

    january_response = await async_client.get(
        "/api/v1/reports/site-daily-labor",
        params={"project_id": str(project.id), "month": "2026-01"},
    )
    assert january_response.status_code == 200
    assert january_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    january_wb = load_workbook(io.BytesIO(january_response.content))
    january_ws = january_wb["명세서"]
    january_worker_row = next(
        row for row in january_ws.iter_rows(values_only=True) if len(row) > 1 and row[1] == "홍길동"
    )
    assert january_worker_row[36] == 2

    february_response = await async_client.get(
        "/api/v1/reports/site-daily-labor",
        params={"project_id": str(project.id), "month": "2026-02"},
    )
    assert february_response.status_code == 200

    february_wb = load_workbook(io.BytesIO(february_response.content))
    february_ws = february_wb["명세서"]
    february_worker_row = next(
        row for row in february_ws.iter_rows(values_only=True) if len(row) > 1 and row[1] == "홍길동"
    )
    assert february_worker_row[36] == 1


@pytest.mark.asyncio
async def test_welfare_form_download_blocks_when_code_mapping_missing(async_client: AsyncClient, labor_api_context):
    db = labor_api_context["db"]
    project, worker = await _seed_project_and_worker(labor_api_context, job_type_code="")

    db.add(
        WorkRecord(
            worker_id=worker.id,
            project_id=project.id,
            work_date=date(2026, 1, 20),
            man_days=Decimal("1"),
        )
    )
    await db.commit()

    response = await async_client.get(
        "/api/v1/reports/welfare-form",
        params={"project_id": str(project.id), "month": "2026-01"},
    )
    assert response.status_code == 422

    detail = response.json()["detail"]
    assert "missing" in detail
    assert detail["missing"][0]["worker_name"] == "홍길동"
    assert "job_type_code" in detail["missing"][0]["missing_fields"]


@pytest.mark.asyncio
async def test_admin_rates_returns_effective_version_by_as_of(async_client: AsyncClient, labor_api_context):
    db = labor_api_context["db"]
    org = labor_api_context["org"]

    db.add(
        InsuranceRate(
            organization_id=org.id,
            effective_year=2025,
            health_insurance_rate=Decimal("0.03495"),
        )
    )
    db.add(
        InsuranceRate(
            organization_id=org.id,
            effective_year=2026,
            health_insurance_rate=Decimal("0.03595"),
        )
    )
    await db.commit()

    response_2025 = await async_client.get("/api/v1/admin/rates", params={"as_of": "2025-12-31"})
    assert response_2025.status_code == 200
    assert response_2025.json()["data"]["active_rate"]["effective_year"] == 2025

    response_2026 = await async_client.get("/api/v1/admin/rates", params={"as_of": "2026-02-01"})
    assert response_2026.status_code == 200
    assert response_2026.json()["data"]["active_rate"]["effective_year"] == 2026


@pytest.mark.asyncio
async def test_labor_codebook_api_returns_expected_keys(async_client: AsyncClient, labor_api_context):
    response = await async_client.get("/api/v1/labor/codebook")
    assert response.status_code == 200

    payload = response.json()["data"]
    assert payload["version"]
    assert "706" in payload["job_type_codes"]
    assert "100" in payload["nationality_codes"]
    assert "E-9" in payload["visa_status_codes"]


@pytest.mark.asyncio
async def test_work_record_batch_is_blocked_when_required_documents_missing(async_client: AsyncClient, labor_api_context):
    project, worker = await _seed_project_and_worker(
        labor_api_context,
        has_id_card=False,
        has_safety_cert=False,
    )

    response = await async_client.post(
        "/api/v1/labor/work-records/batch",
        json={
            "records": [
                {
                    "worker_id": str(worker.id),
                    "project_id": str(project.id),
                    "work_date": "2026-01-20",
                    "man_days": "1.0",
                }
            ]
        },
    )
    assert response.status_code == 422

    detail = response.json()["detail"]
    assert detail["blocked_workers"][0]["worker_name"] == "홍길동"
    assert "id_card" in detail["blocked_workers"][0]["missing_requirements"]
    assert "safety_cert" in detail["blocked_workers"][0]["missing_requirements"]
    assert "consent_record" in detail["blocked_workers"][0]["missing_requirements"]


@pytest.mark.asyncio
async def test_work_record_batch_succeeds_when_gate_requirements_are_met(async_client: AsyncClient, labor_api_context):
    db = labor_api_context["db"]
    org = labor_api_context["org"]
    project, worker = await _seed_project_and_worker(
        labor_api_context,
        has_id_card=True,
        has_safety_cert=True,
        invite_token="invite-token-ok",
    )

    db.add(
        ConsentRecord(
            invite_token="invite-token-ok",
            consent_type="privacy_collection",
            consented=True,
            organization_id=org.id,
        )
    )
    await db.commit()

    response = await async_client.post(
        "/api/v1/labor/work-records/batch",
        json={
            "records": [
                {
                    "worker_id": str(worker.id),
                    "project_id": str(project.id),
                    "work_date": "2026-01-20",
                    "man_days": "1.0",
                }
            ]
        },
    )
    assert response.status_code == 200
    assert response.json()["data"]["updated_count"] == 1


@pytest.mark.asyncio
async def test_consent_and_notification_domains_write_activity_logs(async_client: AsyncClient, labor_api_context):
    db = labor_api_context["db"]
    user = labor_api_context["user"]
    org = labor_api_context["org"]

    role_value = user.role.value if hasattr(user.role, "value") else str(user.role)
    token = create_access_token(
        subject=str(user.id),
        role=role_value,
        org_id=str(org.id),
    )

    consent_response = await async_client.post(
        "/api/v1/consent/records",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "records": [
                {
                    "consent_type": "privacy_collection",
                    "consented": True,
                    "consent_version": "v1.0",
                }
            ]
        },
    )
    assert consent_response.status_code == 201

    invitation_response = await async_client.post(
        "/api/v1/invitations",
        json={"phone": "010-9999-1111", "name": "초대대상", "role": "site_manager"},
    )
    assert invitation_response.status_code == 201

    notification_response = await async_client.post(
        "/api/v1/notifications/alimtalk",
        json={
            "phone": "010-9999-1111",
            "template_code": "USER_INVITE",
            "variables": {"name": "초대대상", "invite_url": "/accept-invite/mock"},
        },
    )
    assert notification_response.status_code == 200

    logs = (
        await db.execute(
            select(ActivityLog)
            .where(ActivityLog.user_id == user.id)
            .order_by(ActivityLog.created_at.asc())
        )
    ).scalars().all()
    actions = [log.action for log in logs]

    assert "consent_recorded" in actions
    assert "invitation_create" in actions
    assert "notification_send_success" in actions
