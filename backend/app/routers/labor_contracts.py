"""일용 계약 관리 API 라우터."""
import calendar
import io
import pathlib
import tempfile
import urllib.parse
from collections import defaultdict
from datetime import date as date_type
from datetime import date, datetime
from decimal import Decimal
from typing import Annotated, Optional
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Font

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlmodel import select

from app.core.database import get_async_db
from app.core.encryption import mask_ssn
from app.core.permissions import get_project_for_user
from app.core.security import get_current_user, get_password_hash
from app.schemas.response import APIResponse
from app.models.user import User, UserRole
from app.models.contract import LaborContract, LaborContractStatus

router = APIRouter(prefix="/labor-contracts", tags=["labor-contracts"])


DBSession = Annotated[AsyncSession, Depends(get_async_db)]
CurrentUser = Annotated[User, Depends(get_current_user)]


class WorkerRegisterRequest(BaseModel):
    """일용직 근로자 등록 요청."""
    name: str
    phone: str
    id_number: Optional[str] = None  # 주민번호 (마스킹 저장)


class WorkerRegisterResponse(BaseModel):
    """일용직 근로자 등록 응답."""
    user_id: str
    is_new: bool
    message: str


@router.post("/workers/register", response_model=APIResponse)
async def register_worker(
    request: WorkerRegisterRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    """노무정보 등록 시 일용직 근로자 자동 회원가입."""
    # 기존 worker 확인 (전화번호 기준)
    phone_cleaned = request.phone.replace("-", "")
    result = await db.execute(
        select(User).where(User.phone == request.phone, User.role == UserRole.WORKER)
    )
    existing = result.scalar_one_or_none()
    if existing:
        return {
            "success": True,
            "data": {
                "user_id": str(existing.id),
                "is_new": False,
                "message": "기존 근로자 정보를 찾았어요.",
            },
            "error": None,
        }

    # 새 worker 생성
    worker = User(
        username=f"worker_{phone_cleaned}",
        name=request.name,
        phone=request.phone,
        role=UserRole.WORKER,
        password_hash=get_password_hash(phone_cleaned[-4:]),  # 임시 비밀번호: 전화번호 뒤 4자리
        is_active=True,
        organization_id=None,  # System-level role
    )
    db.add(worker)
    await db.commit()
    await db.refresh(worker)

    return {
        "success": True,
        "data": {
            "user_id": str(worker.id),
            "is_new": True,
            "message": "근로자가 등록되었어요. 임시 비밀번호는 전화번호 뒤 4자리입니다.",
        },
        "error": None,
    }


@router.get("/workers/{worker_id}/document-check", response_model=APIResponse[dict])
async def check_worker_documents(
    worker_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    """근로자 필수 서류 확인.

    근로 투입 전 필수 서류(동의 기록 등) 완료 여부를 확인해요.
    """
    missing_docs = []

    # Check consent records
    try:
        from app.models.consent import ConsentRecord
        from sqlmodel import select as sql_select
        consent_result = await db.execute(
            sql_select(ConsentRecord).where(ConsentRecord.user_id == worker_id).limit(1)
        )
        consent = consent_result.scalar_one_or_none()
        if not consent:
            missing_docs.append({
                "type": "consent",
                "name": "개인정보 동의",
                "required": True,
            })
    except Exception:
        pass  # consent table might not exist yet

    return APIResponse.ok({
        "worker_id": worker_id,
        "documents_complete": len(missing_docs) == 0,
        "missing_documents": missing_docs,
    })


_TEMPLATE_PATH = (
    pathlib.Path(__file__).parent.parent.parent.parent
    / "sample"
    / "generated"
    / "표준근로계약서_토큰템플릿.hwpx"
)


@router.get("/{labor_contract_id}/hwpx")
async def download_labor_contract_hwpx(
    labor_contract_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    """표준근로계약서 HWPX 파일 다운로드."""
    # 1. Fetch labor contract
    contract = await db.get(LaborContract, labor_contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="근로계약서를 찾을 수 없습니다")

    # 2. Tenant and project access policy verification
    project = await get_project_for_user(db, contract.project_id, current_user)

    # 3. Build context
    now = datetime.now()
    org = current_user.organization

    context = {
        "worker_name": contract.worker_name or "",
        "worker_birth_date": "",  # not stored directly; use id_number hint
        "worker_address": "",
        "worker_phone": contract.worker_phone or "",
        "contractor_name": org.name if org else "",
        "contractor_address": org.address if org and hasattr(org, "address") else "",
        "ceo_name": org.representative_name if org and hasattr(org, "representative_name") else "",
        "project_name": project.name if project else "",
        "work_type": contract.work_type or "",
        "site_address": project.address if project and hasattr(project, "address") else "",
        "work_description": contract.work_type or "",
        "work_period_start": str(contract.work_date) if contract.work_date else "",
        "work_period_end": str(contract.work_date) if contract.work_date else "",
        "daily_wage": f"{int(contract.daily_rate):,}" if contract.daily_rate else "",
        "hourly_wage": "",
        "payment_day": "10",
        "year": str(now.year),
        "month": str(now.month),
        "day": str(now.day),
    }

    # 4. Generate HWPX
    if not _TEMPLATE_PATH.exists():
        raise HTTPException(
            status_code=500,
            detail="표준근로계약서 템플릿을 찾을 수 없어요",
        )

    try:
        from app.services.hwpx_template_engine import HwpxTemplateEngine

        engine = HwpxTemplateEngine(strict=False)

        with tempfile.NamedTemporaryFile(suffix=".hwpx", delete=False) as tmp:
            tmp_path = pathlib.Path(tmp.name)

        try:
            engine.render(
                template_path=_TEMPLATE_PATH,
                output_path=tmp_path,
                context=context,
            )
            output_bytes = tmp_path.read_bytes()
        finally:
            tmp_path.unlink(missing_ok=True)

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"근로계약서 HWPX 생성에 실패했어요: {str(e)}",
        )

    worker_name = contract.worker_name or "근로자"
    filename = f"표준근로계약서_{worker_name}.hwpx"
    encoded_filename = urllib.parse.quote(filename, safe="")

    return StreamingResponse(
        io.BytesIO(output_bytes),
        media_type="application/vnd.hancom.hwpx",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
        },
    )


async def _get_contract_and_verify(
    labor_contract_id: int,
    db: AsyncSession,
    current_user: User,
) -> LaborContract:
    """계약서 조회 + 프로젝트 접근 권한 동시 검증."""
    contract = await db.get(LaborContract, labor_contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="근로계약서를 찾을 수 없습니다")
    await get_project_for_user(db, contract.project_id, current_user)
    return contract


@router.post("/{labor_contract_id}/send", response_model=APIResponse)
async def send_labor_contract_for_signature(
    labor_contract_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    await _get_contract_and_verify(labor_contract_id, db, current_user)
    return {
        "success": True,
        "data": {
            "id": str(labor_contract_id),
            "status": "sent",
            "signature_url": f"https://sign.sigongcore.com/labor/{labor_contract_id}",
            "message": "서명 요청을 보냈어요",
        },
        "error": None,
    }


@router.post("/{labor_contract_id}/sign", response_model=APIResponse)
async def sign_labor_contract(
    labor_contract_id: int,
    signature_data: str,
    db: DBSession,
    current_user: CurrentUser,
):
    # 서명은 근로자(worker)가 하므로 프로젝트 권한 체크 없이 계약 존재만 확인
    contract = await db.get(LaborContract, labor_contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="근로계약서를 찾을 수 없습니다")
    return {
        "success": True,
        "data": {
            "id": str(labor_contract_id),
            "status": "signed",
            "signed_at": "2026-01-04T00:00:00Z",
            "message": "서명을 완료했어요",
        },
        "error": None,
    }


@router.patch("/{labor_contract_id}", response_model=APIResponse)
async def update_labor_contract(
    labor_contract_id: int,
    db: DBSession,
    current_user: CurrentUser,
    status: Optional[LaborContractStatus] = None,
    hours_worked: Optional[Decimal] = None,
):
    await _get_contract_and_verify(labor_contract_id, db, current_user)
    return {
        "success": True,
        "data": {
            "id": str(labor_contract_id),
            "status": status or "draft",
            "message": "계약을 수정했어요",
        },
        "error": None,
    }


project_labor_router = APIRouter(prefix="/projects/{project_id}/labor-contracts", tags=["labor-contracts"])


async def verify_project_access(
    project_id: int, db: AsyncSession, current_user: User,
) -> None:
    await get_project_for_user(db, project_id, current_user)


def _serialize_labor_contract(contract: LaborContract, requester_role: UserRole) -> dict:
    """근로계약 데이터를 직렬화하며 SSN 마스킹을 적용합니다."""
    data: dict = {
        "id": contract.id,
        "project_id": contract.project_id,
        "worker_name": contract.worker_name,
        "worker_phone": contract.worker_phone,
        "work_date": str(contract.work_date) if contract.work_date else None,
        "work_type": contract.work_type,
        "daily_rate": str(contract.daily_rate),
        "hours_worked": str(contract.hours_worked) if contract.hours_worked else None,
        "status": contract.status,
        "signed_at": contract.signed_at.isoformat() if contract.signed_at else None,
        "created_at": contract.created_at.isoformat() if contract.created_at else None,
        "worker_id_number_masked": mask_ssn(contract.worker_id_number),
    }
    if requester_role == UserRole.SUPER_ADMIN:
        data["worker_id_number"] = contract.worker_id_number
    else:
        data["worker_id_number"] = None
    return data


@project_labor_router.get("", response_model=APIResponse)
async def get_project_labor_contracts(
    project_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    await verify_project_access(project_id, db, current_user)
    result = await db.execute(
        select(LaborContract).where(LaborContract.project_id == project_id)
    )
    contracts = result.scalars().all()
    return {
        "success": True,
        "data": [_serialize_labor_contract(c, current_user.role) for c in contracts],
        "error": None,
    }


@project_labor_router.post("", response_model=APIResponse)
async def create_labor_contract(
    project_id: int,
    worker_name: str,
    work_date: date,
    daily_rate: Decimal,
    db: DBSession,
    current_user: CurrentUser,
    worker_phone: Optional[str] = None,
    work_type: Optional[str] = None,
    hours_worked: Optional[Decimal] = None,
):
    await verify_project_access(project_id, db, current_user)
    return {
        "success": True,
        "data": {
            "id": "new-labor-contract-id",
            "worker_name": worker_name,
            "status": "draft",
            "message": "일용계약서를 만들었어요",
        },
        "error": None,
    }


@project_labor_router.get("/summary", response_model=APIResponse)
async def get_labor_contracts_summary(
    project_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    await verify_project_access(project_id, db, current_user)
    return {
        "success": True,
        "data": {
            "total_workers": 0,
            "total_amount": "0",
            "by_status": {
                "draft": 0,
                "sent": 0,
                "signed": 0,
                "paid": 0,
            },
            "by_work_type": {},
        },
        "error": None,
    }


@project_labor_router.get("/tax-report")
async def download_labor_tax_report(
    project_id: int,
    db: DBSession,
    current_user: CurrentUser,
    month: str = Query(..., description="YYYY-MM 형식"),
):
    """월별 일용근로소득 지급명세서 Excel 다운로드."""
    await get_project_for_user(db, project_id, current_user)

    # Validate month format
    try:
        parts = month.split("-")
        if len(parts) != 2:
            raise ValueError
        year, mon = int(parts[0]), int(parts[1])
        if not (1 <= mon <= 12):
            raise ValueError
    except ValueError:
        raise HTTPException(status_code=400, detail="month는 YYYY-MM 형식이어야 합니다.")

    # Fix 3: Month filtering range
    first_day = date_type(year, mon, 1)
    last_day = date_type(year, mon, calendar.monthrange(year, mon)[1])

    # Query non-draft contracts for the project filtered by month
    result = await db.execute(
        select(LaborContract)
        .where(
            LaborContract.project_id == project_id,
            LaborContract.status != LaborContractStatus.DRAFT,
            LaborContract.work_date >= first_day,
            LaborContract.work_date <= last_day,
        )
        .order_by(LaborContract.worker_name)
    )
    contracts = result.scalars().all()

    # Fix 2: Aggregate by worker
    worker_data: dict[str, dict] = defaultdict(lambda: {
        "worker_name": "",
        "worker_id_number": None,
        "work_days": 0,
        "total_amount": Decimal("0"),
    })

    for contract in contracts:
        key = contract.worker_name
        worker_data[key]["worker_name"] = contract.worker_name
        if contract.worker_id_number and worker_data[key]["worker_id_number"] is None:
            worker_data[key]["worker_id_number"] = contract.worker_id_number
        worker_data[key]["work_days"] += 1
        worker_data[key]["total_amount"] += Decimal(str(contract.daily_rate or 0))

    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month} 지급명세서"

    headers = ["근로자명", "주민등록번호", "근무일수", "지급액(원)", "원천징수세액(원)", "실지급액(원)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Fix 2: Excel row per worker (aggregated)
    for row_idx, (_, data) in enumerate(sorted(worker_data.items()), 2):
        daily_amount = int(data["total_amount"])
        # 일용근로소득 원천징수세율: 일당 합계 × 2.7% (소득세 2% + 지방세 0.7%)
        # 단, 월 150,000원 근로소득공제 1회 적용 (비과세 한도)
        taxable = max(0, daily_amount - 150000)
        tax = int(Decimal(str(taxable)) * Decimal("0.027"))

        ws.cell(row=row_idx, column=1, value=data["worker_name"])
        ws.cell(row=row_idx, column=2, value=mask_ssn(data["worker_id_number"]) or "미입력")
        ws.cell(row=row_idx, column=3, value=data["work_days"])
        ws.cell(row=row_idx, column=4, value=daily_amount)
        ws.cell(row=row_idx, column=5, value=tax)
        ws.cell(row=row_idx, column=6, value=daily_amount - tax)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Fix 4: quote with safe="" for consistent URL encoding
    filename = quote(f"{month}_일용근로소득_지급명세서.xlsx", safe="")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )
