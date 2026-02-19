"""Auxiliary operations router for frontend API parity."""
from __future__ import annotations

import io
import hashlib
import random
import secrets
import tempfile
from datetime import date, datetime, timedelta
from decimal import ROUND_DOWN, Decimal
from pathlib import Path
from typing import Annotated, Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlmodel import select

from app.core.config import settings
from app.core.database import get_async_db
from app.core.exceptions import NotFoundException
from app.core.permissions import (
    ROLE_COMPANY_ADMIN,
    ROLE_SITE_MANAGER,
    ROLE_SUPER_ADMIN,
    ROLE_WORKER,
    ensure_can_create_invitation,
    get_project_for_user,
)
from app.core.security import get_current_user, get_password_hash, verify_password
from app.models.billing import Payment, PaymentStatus, Subscription, SubscriptionPlan, SubscriptionStatus
from app.models.consent import ConsentRecord
from app.models.contract import LaborContract, LaborContractStatus
from app.models.diagnosis import AIDiagnosis
from app.models.operations import (
    AccountRequest,
    ActivityLog,
    AppNotification,
    DailyReport,
    DailyWorker,
    InsuranceRate,
    Invitation,
    InvitationStatus,
    MaterialOrder,
    MaterialOrderItem,
    MaterialOrderStatus,
    NotificationType,
    Partner,
    PartnerStatus,
    Paystub,
    PaystubItem,
    PaystubStatus,
    ProjectAccessPolicy,
    UtilityDocStatus,
    UtilityItem,
    UtilityStatus,
    UtilityTimeline,
    UserNotificationPrefs,
    WorkRecord,
    WorkerAccessRequest,
    WorkerDocument,
    WorkerDocumentModerationLog,
)
from app.models.project import Project, ProjectStatus, SiteVisit
from app.models.pricebook import CatalogItem, CatalogItemPrice
from app.models.user import Organization, User, UserRole
from app.schemas.response import APIResponse, PaginatedResponse
from app.services.audit_log import write_activity_log
from app.services.labor_codebook import (
    get_labor_codebook_payload,
    is_valid_job_type_code,
    is_valid_nationality_code,
    is_valid_visa_status,
    normalize_code,
)
from app.services.storage import storage_service

router = APIRouter()

DBSession = Annotated[AsyncSession, Depends(get_async_db)]
CurrentUser = Annotated[User, Depends(get_current_user)]

_WEATHER_LABELS: dict[str, str] = {
    "sunny": "맑음",
    "cloudy": "흐림",
    "rain": "비",
    "rainy": "비",
    "snow": "눈",
}
_SAMPLE_ROOT = Path(__file__).parent.parent.parent.parent / "sample"
_DAILY_REPORT_TEMPLATE_CANDIDATES: tuple[Path, ...] = (
    _SAMPLE_ROOT / "generated" / "공사일지_토큰템플릿.hwpx",
    _SAMPLE_ROOT / "7. 공사일지" / "공사일지_토큰템플릿.hwpx",
)

_WORKER_DOCUMENT_TYPE_ID_CARD = "id_card"
_WORKER_DOCUMENT_TYPE_SAFETY_CERT = "safety_cert"
_WORKER_DOCUMENT_ALLOWED_TYPES = {
    _WORKER_DOCUMENT_TYPE_ID_CARD,
    _WORKER_DOCUMENT_TYPE_SAFETY_CERT,
}
_WORKER_DOCUMENT_ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".pdf"}
_WORKER_DOCUMENT_ALLOWED_MIME_TYPES = {
    "image/jpeg",
    "image/png",
    "application/pdf",
}
_WORKER_DOCUMENT_LABELS = {
    _WORKER_DOCUMENT_TYPE_ID_CARD: "신분증",
    _WORKER_DOCUMENT_TYPE_SAFETY_CERT: "안전교육 이수증",
}
_WORKER_DOCUMENT_REVIEW_STATUSES = {"pending_review", "approved", "rejected", "quarantined"}
_WORKER_DOCUMENT_REVIEW_ACTIONS = {"approve", "reject", "quarantine", "request_reupload"}
_WORKER_DOCUMENT_CONTROL_ACTIONS = {"block", "unblock"}

_MATERIAL_ORDER_FINANCE_STATUSES = {
    "invoice_received",
    "payment_completed",
    "closed",
}
_MATERIAL_ORDER_SITE_MANAGER_ALLOWED_TARGETS = {
    "requested",
    "delivered",
}


def _role_value(user: User) -> str:
    role = user.role
    return role.value if hasattr(role, "value") else str(role)


def _require_org_id(user: User) -> int:
    if user.organization_id is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="조직 정보가 필요한 요청이에요.",
        )
    return user.organization_id


def _require_non_super_admin_for_labor_actions(user: User) -> None:
    if _role_value(user) == ROLE_SUPER_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="회원기업 노무관리에서만 가능한 작업이에요.",
        )


def _require_worker_document_upload_role(user: User) -> None:
    role = _role_value(user)
    if role not in {ROLE_COMPANY_ADMIN, ROLE_SITE_MANAGER, ROLE_SUPER_ADMIN}:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="근로자 서류 업로드 권한이 없어요.",
        )


def _require_worker_document_review_role(user: User) -> None:
    role = _role_value(user)
    if role not in {ROLE_COMPANY_ADMIN, ROLE_SUPER_ADMIN}:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="근로자 서류 검토/통제 권한이 없어요.",
        )


def _normalize_worker_document_type(raw: Optional[str]) -> str:
    value = (raw or "").strip().lower()
    if value in _WORKER_DOCUMENT_ALLOWED_TYPES:
        return value
    if value == "doc_1":
        return _WORKER_DOCUMENT_TYPE_ID_CARD
    if value == "doc_2":
        return _WORKER_DOCUMENT_TYPE_SAFETY_CERT
    return value or "unknown"


def _require_valid_worker_document_type(document_type: str) -> str:
    normalized = _normalize_worker_document_type(document_type)
    if normalized not in _WORKER_DOCUMENT_ALLOWED_TYPES:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="지원하지 않는 근로자 서류 유형이에요.",
        )
    return normalized


def _worker_document_label(document_type: str) -> str:
    return _WORKER_DOCUMENT_LABELS.get(document_type, document_type)


async def _get_daily_worker_for_document_ops(
    db: AsyncSession,
    worker_id: int,
    current_user: User,
) -> DailyWorker:
    query = select(DailyWorker).where(DailyWorker.id == worker_id)
    if current_user.organization_id is not None:
        query = query.where(DailyWorker.organization_id == current_user.organization_id)

    worker = (await db.execute(query)).scalar_one_or_none()
    if not worker:
        raise NotFoundException("daily_worker", worker_id)
    return worker


def _validate_worker_document_file(file: UploadFile) -> tuple[str, str]:
    ext = Path(file.filename or "").suffix.lower()
    if ext not in _WORKER_DOCUMENT_ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="PDF 또는 JPG/PNG 파일만 업로드할 수 있어요.",
        )

    mime_type = (file.content_type or "").lower()
    if mime_type and mime_type not in _WORKER_DOCUMENT_ALLOWED_MIME_TYPES:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="지원하지 않는 파일 형식이에요.",
        )

    if not storage_service.validate_file_size(file):
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"파일은 {settings.max_upload_size_mb}MB 이하로 올려주세요.",
        )

    return ext, mime_type


def _detect_worker_document_anomaly_flags(
    *,
    content: bytes,
    ext: str,
    mime_type: str,
    duplicate_hash_exists: bool,
) -> list[str]:
    flags: list[str] = []
    if len(content) < 512:
        flags.append("small_file")

    if ext == ".pdf" and not content.startswith(b"%PDF"):
        flags.append("pdf_signature_mismatch")
    if ext in {".jpg", ".jpeg"} and not content.startswith(b"\xff\xd8\xff"):
        flags.append("jpeg_signature_mismatch")
    if ext == ".png" and not content.startswith(b"\x89PNG\r\n\x1a\n"):
        flags.append("png_signature_mismatch")

    if duplicate_hash_exists:
        flags.append("duplicate_hash")

    if mime_type == "application/pdf" and ext != ".pdf":
        flags.append("mime_extension_mismatch")
    if mime_type.startswith("image/") and ext == ".pdf":
        flags.append("mime_extension_mismatch")

    return flags


def _serialize_worker_document(document: WorkerDocument) -> dict:
    document_type = _normalize_worker_document_type(
        getattr(document, "document_type", None) or document.document_id
    )
    review_status = (document.review_status or "").strip().lower()
    if review_status not in _WORKER_DOCUMENT_REVIEW_STATUSES:
        review_status = "pending_review"

    return {
        "id": str(document.id),
        "worker_id": str(document.worker_id),
        "organization_id": str(document.organization_id) if document.organization_id is not None else None,
        "document_id": document.document_id,
        "document_type": document_type,
        "document_name": _worker_document_label(document_type),
        "name": document.name,
        "status": document.status,
        "storage_path": document.storage_path,
        "original_filename": document.original_filename,
        "mime_type": document.mime_type,
        "file_size_bytes": int(document.file_size_bytes or 0),
        "file_hash_sha256": document.file_hash_sha256,
        "review_status": review_status,
        "review_reason": document.review_reason,
        "reviewed_by_user_id": str(document.reviewed_by_user_id) if document.reviewed_by_user_id else None,
        "reviewed_at": _to_iso(document.reviewed_at),
        "anomaly_flags": list(document.anomaly_flags or []),
        "uploaded_at": _to_iso(document.uploaded_at),
    }


async def _sync_daily_worker_document_flags(
    db: AsyncSession,
    worker: DailyWorker,
) -> None:
    docs = (
        await db.execute(
            select(WorkerDocument).where(WorkerDocument.worker_id == worker.id)
        )
    ).scalars().all()

    has_id_card = any(
        _normalize_worker_document_type(doc.document_type or doc.document_id) == _WORKER_DOCUMENT_TYPE_ID_CARD
        and doc.review_status == "approved"
        and bool(doc.storage_path)
        for doc in docs
    )
    has_safety_cert = any(
        _normalize_worker_document_type(doc.document_type or doc.document_id) == _WORKER_DOCUMENT_TYPE_SAFETY_CERT
        and doc.review_status == "approved"
        and bool(doc.storage_path)
        for doc in docs
    )

    worker.has_id_card = has_id_card
    worker.has_safety_cert = has_safety_cert
    if not (has_id_card and has_safety_cert):
        if worker.registration_status == "registered":
            worker.registration_status = "pending_docs"
    elif worker.registration_status in {"pending_docs", "invited"}:
        worker.registration_status = "registered"
    worker.updated_at = datetime.utcnow()
    await db.flush()


async def _write_worker_document_moderation_log(
    db: AsyncSession,
    *,
    organization_id: int,
    worker_id: int,
    actor_user_id: int,
    action: str,
    reason: Optional[str] = None,
    worker_document_id: Optional[int] = None,
    payload: Optional[dict] = None,
) -> None:
    db.add(
        WorkerDocumentModerationLog(
            organization_id=organization_id,
            worker_id=worker_id,
            worker_document_id=worker_document_id,
            actor_user_id=actor_user_id,
            action=action,
            reason=reason,
            payload_json=payload or {},
        )
    )
    await db.flush()


async def _get_project_for_user(
    db: AsyncSession,
    project_id: int,
    user: User,
) -> Project:
    return await get_project_for_user(db, project_id, user)


async def _get_daily_report_for_project(
    db: AsyncSession,
    project_id: int,
    report_id: int,
) -> DailyReport:
    result = await db.execute(
        select(DailyReport)
        .where(DailyReport.project_id == project_id)
        .where(DailyReport.id == report_id)
    )
    report = result.scalar_one_or_none()
    if not report:
        raise NotFoundException("daily_report", report_id)
    return report


def _to_int(value: int | str) -> int:
    if isinstance(value, int):
        return value
    raw = str(value)
    if raw.isdigit():
        return int(raw)
    digits = "".join(ch for ch in raw if ch.isdigit())
    if digits:
        return int(digits)
    raise ValueError(f"Cannot parse integer from value: {value}")


def _to_iso(dt: datetime | None) -> str | None:
    return dt.isoformat() if dt else None


def _safe_decimal_to_int(value: Decimal | int | float | None) -> int:
    if value is None:
        return 0
    if isinstance(value, Decimal):
        return int(value)
    return int(value)


def _normalize_material_order_status_value(raw_status: MaterialOrderStatus | str) -> str:
    value = raw_status.value if hasattr(raw_status, "value") else str(raw_status)
    if value == MaterialOrderStatus.CONFIRMED.value:
        return MaterialOrderStatus.INVOICE_RECEIVED.value
    return value


def _material_order_allowed_transitions() -> dict[str, set[str]]:
    return {
        MaterialOrderStatus.DRAFT.value: {
            MaterialOrderStatus.REQUESTED.value,
            MaterialOrderStatus.CANCELLED.value,
        },
        MaterialOrderStatus.REQUESTED.value: {
            MaterialOrderStatus.INVOICE_RECEIVED.value,
            MaterialOrderStatus.CANCELLED.value,
        },
        MaterialOrderStatus.INVOICE_RECEIVED.value: {
            MaterialOrderStatus.PAYMENT_COMPLETED.value,
            MaterialOrderStatus.CANCELLED.value,
        },
        MaterialOrderStatus.PAYMENT_COMPLETED.value: {
            MaterialOrderStatus.SHIPPED.value,
            MaterialOrderStatus.CANCELLED.value,
        },
        MaterialOrderStatus.SHIPPED.value: {
            MaterialOrderStatus.DELIVERED.value,
            MaterialOrderStatus.CANCELLED.value,
        },
        MaterialOrderStatus.DELIVERED.value: {
            MaterialOrderStatus.CLOSED.value,
            MaterialOrderStatus.CANCELLED.value,
        },
        MaterialOrderStatus.CLOSED.value: set(),
        MaterialOrderStatus.CANCELLED.value: set(),
    }


def _require_material_order_status_permission(
    *,
    user: User,
    target_status: str,
) -> None:
    role = _role_value(user)

    if target_status in _MATERIAL_ORDER_FINANCE_STATUSES and role not in {
        ROLE_COMPANY_ADMIN,
        ROLE_SUPER_ADMIN,
    }:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="금액/정산 관련 발주 상태는 대표자 또는 최고관리자만 변경할 수 있어요.",
        )

    if role == ROLE_SITE_MANAGER and target_status not in _MATERIAL_ORDER_SITE_MANAGER_ALLOWED_TARGETS:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="현장소장은 발주 요청/수령 확인 상태만 변경할 수 있어요.",
        )


def _serialize_material_order_item(item: MaterialOrderItem) -> dict:
    return {
        "id": str(item.id),
        "catalog_item_id": str(item.catalog_item_id) if item.catalog_item_id else None,
        "pricebook_revision_id": str(item.pricebook_revision_id) if item.pricebook_revision_id else None,
        "price_source": item.price_source,
        "override_reason": item.override_reason,
        "description": item.description,
        "specification": item.specification,
        "unit": item.unit,
        "quantity": float(item.quantity),
        "unit_price": float(item.unit_price),
        "amount": float(item.amount),
    }


def _serialize_material_order(order: MaterialOrder, items: list[MaterialOrderItem]) -> dict:
    normalized_status = _normalize_material_order_status_value(order.status)
    return {
        "id": str(order.id),
        "project_id": str(order.project_id),
        "order_number": order.order_number,
        "status": normalized_status,
        "total_amount": order.total_amount,
        "vendor_id": str(order.vendor_id) if order.vendor_id else None,
        "invoice_number": order.invoice_number,
        "invoice_amount": order.invoice_amount,
        "invoice_file_url": order.invoice_file_url,
        "requested_at": _to_iso(order.requested_at),
        "confirmed_at": _to_iso(order.confirmed_at),
        "payment_at": _to_iso(order.payment_at),
        "shipped_at": _to_iso(order.shipped_at),
        "delivered_at": _to_iso(order.delivered_at),
        "received_at": _to_iso(order.received_at),
        "received_by_user_id": str(order.received_by_user_id) if order.received_by_user_id else None,
        "closed_at": _to_iso(order.closed_at),
        "notes": order.notes,
        "created_at": order.created_at.isoformat(),
        "updated_at": order.updated_at.isoformat(),
        "items": [_serialize_material_order_item(item) for item in items],
    }


def _serialize_daily_report(report: DailyReport) -> dict:
    photos = report.photos or []
    return {
        "id": str(report.id),
        "project_id": str(report.project_id),
        "work_date": report.work_date.isoformat(),
        "weather": report.weather,
        "temperature": report.temperature,
        "work_description": report.work_description,
        "tomorrow_plan": report.tomorrow_plan,
        "photos": photos,
        "photo_count": len(photos),
        "created_at": report.created_at.isoformat(),
    }


def _format_temperature_display(value: str | None) -> str:
    if not value:
        return ""
    text = value.strip()
    if not text:
        return ""
    if "℃" in text or "°C" in text:
        return text
    return f"{text}℃"


def _resolve_daily_report_template_path() -> Path | None:
    for template_path in _DAILY_REPORT_TEMPLATE_CANDIDATES:
        if template_path.exists():
            return template_path

    sample_daily_report_dir = _SAMPLE_ROOT / "7. 공사일지"
    if sample_daily_report_dir.exists():
        candidates = sorted(sample_daily_report_dir.glob("*.hwpx"))
        if candidates:
            return candidates[0]
    return None


def _build_daily_report_hwpx_context(project: Project, report: DailyReport) -> dict:
    weather_label = _WEATHER_LABELS.get((report.weather or "").lower(), report.weather or "")
    temperature_display = _format_temperature_display(report.temperature)
    photos = report.photos or []
    return {
        "project_name": project.name or "",
        "site_address": project.address or "",
        "work_date": report.work_date.isoformat(),
        "work_date_kor": report.work_date.strftime("%Y년 %m월 %d일"),
        "weather": weather_label,
        "weather_code": report.weather or "",
        "temperature": temperature_display,
        "temperature_raw": (report.temperature or "").strip(),
        "work_description": report.work_description or "",
        "tomorrow_plan": report.tomorrow_plan or "",
        "photo_count": str(len(photos)),
        "photo_count_number": len(photos),
        "photos": [
            {
                "index": idx + 1,
                "path": path,
                "item": path,
            }
            for idx, path in enumerate(photos)
        ],
        "photos_summary": ", ".join(photos),
    }


async def _log_activity(
    db: AsyncSession,
    user_id: int,
    action: str,
    description: str,
) -> None:
    await write_activity_log(
        db,
        user_id=user_id,
        action=action,
        description=description,
        ip_address="0.0.0.0",
        device_info="web",
    )


def _normalized_or_none(value: Optional[str]) -> Optional[str]:
    normalized = normalize_code(value)
    return normalized or None


def _normalize_phone_digits(value: Optional[str]) -> str:
    if not value:
        return ""
    return "".join(ch for ch in value if ch.isdigit())


def _validate_daily_worker_payload(payload: "DailyWorkerUpsertRequest") -> None:
    if not settings.labor_code_validation_enabled:
        return

    errors: list[dict] = []
    job_type_code = normalize_code(payload.job_type_code)
    if not job_type_code:
        errors.append({"field": "job_type_code", "reason": "required"})
    elif not is_valid_job_type_code(job_type_code):
        errors.append({"field": "job_type_code", "reason": "invalid", "value": job_type_code})

    if payload.is_foreign:
        nationality_code = normalize_code(payload.nationality_code)
        visa_status = normalize_code(payload.visa_status)

        if not nationality_code:
            errors.append({"field": "nationality_code", "reason": "required"})
        elif not is_valid_nationality_code(nationality_code):
            errors.append({"field": "nationality_code", "reason": "invalid", "value": nationality_code})

        if not visa_status:
            errors.append({"field": "visa_status", "reason": "required"})
        elif not is_valid_visa_status(visa_status):
            errors.append({"field": "visa_status", "reason": "invalid", "value": visa_status})

    if errors:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "근로자 코드 매핑 값이 올바르지 않아요.",
                "errors": errors,
            },
        )


async def _has_consent_record_for_worker(db: AsyncSession, worker: DailyWorker) -> bool:
    if worker.invite_token:
        invite_token_result = await db.execute(
            select(ConsentRecord.id)
            .where(ConsentRecord.invite_token == worker.invite_token)
            .where(ConsentRecord.consented == True)
            .limit(1)
        )
        if invite_token_result.scalar_one_or_none():
            return True

    normalized_phone = _normalize_phone_digits(worker.phone)
    if not normalized_phone:
        return False

    user_ids_result = await db.execute(
        select(User.id)
        .where(User.organization_id == worker.organization_id)
        .where(func.replace(func.coalesce(User.phone, ""), "-", "") == normalized_phone)
    )
    user_ids = [row[0] for row in user_ids_result.all()]
    if not user_ids:
        return False

    consent_result = await db.execute(
        select(ConsentRecord.id)
        .where(ConsentRecord.user_id.in_(user_ids))
        .where(ConsentRecord.consented == True)
        .limit(1)
    )
    return consent_result.scalar_one_or_none() is not None


async def _get_or_create_notification_prefs(
    db: AsyncSession,
    user_id: int,
) -> UserNotificationPrefs:
    result = await db.execute(select(UserNotificationPrefs).where(UserNotificationPrefs.user_id == user_id))
    prefs = result.scalar_one_or_none()
    if prefs:
        return prefs

    prefs = UserNotificationPrefs(user_id=user_id)
    db.add(prefs)
    await db.flush()
    return prefs


async def _seed_utilities_if_empty(db: AsyncSession, project_id: int) -> None:
    count = (await db.execute(
        select(func.count()).select_from(UtilityItem).where(UtilityItem.project_id == project_id)
    )).scalar() or 0
    if count > 0:
        return

    today = date.today()
    month_str = f"{today.year}-{today.month:02d}"
    types = ["수도", "전기", "가스"]
    for idx, utility_type in enumerate(types):
        due_date = today + timedelta(days=7 + idx * 3)
        item = UtilityItem(
            project_id=project_id,
            type=utility_type,
            month=month_str,
            status=UtilityStatus.PENDING,
            amount=120000 + idx * 30000,
            due_date=due_date,
            doc_status=UtilityDocStatus.PENDING,
        )
        db.add(item)
    db.add(
        UtilityTimeline(
            project_id=project_id,
            date=datetime.utcnow(),
            message="수도광열비 정산 항목을 초기화했어요.",
        )
    )
    await db.flush()


async def _resolve_worker_user_id(
    db: AsyncSession,
    worker_id: str,
    *,
    create_if_missing: bool = False,
) -> int:
    candidate_id: Optional[int] = None

    if worker_id.isdigit():
        candidate_id = int(worker_id)
    elif worker_id.startswith("worker_") and worker_id.split("worker_", 1)[1].isdigit():
        candidate_id = int(worker_id.split("worker_", 1)[1])

    if candidate_id is not None:
        result = await db.execute(select(User).where(User.id == candidate_id))
        user = result.scalar_one_or_none()
        if user:
            return user.id

    result = await db.execute(select(User).where(User.username == worker_id))
    user = result.scalar_one_or_none()
    if user:
        return user.id

    if not create_if_missing:
        raise NotFoundException("worker", worker_id)

    # fallback worker account (legacy compatibility)
    username = worker_id if worker_id.startswith("worker_") else f"worker_{worker_id}"
    fallback = User(
        username=username,
        name="작업자",
        phone=None,
        email=None,
        role=UserRole.WORKER,
        organization_id=None,
        password_hash=get_password_hash("123456"),
        is_active=True,
    )
    db.add(fallback)
    await db.flush()
    return fallback.id


def _ensure_worker_role(current_user: User) -> None:
    if _role_value(current_user) != ROLE_WORKER:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="근로자 계정만 접근할 수 있어요",
        )


async def _require_worker_self(
    db: AsyncSession,
    current_user: User,
    worker_id: str,
) -> int:
    _ensure_worker_role(current_user)
    worker_user_id = await _resolve_worker_user_id(db, worker_id)
    if worker_user_id != int(current_user.id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="본인 정보만 조회할 수 있어요",
        )
    return worker_user_id


def _require_worker_contract_access(contract: LaborContract, current_user: User) -> None:
    _ensure_worker_role(current_user)
    worker_phone = _normalize_phone_digits(contract.worker_phone)
    current_phone = _normalize_phone_digits(current_user.phone)
    if not worker_phone or not current_phone or worker_phone != current_phone:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="본인 계약서만 조회할 수 있어요",
        )


def _require_project_policy_admin(current_user: User) -> None:
    if _role_value(current_user) not in {ROLE_COMPANY_ADMIN, ROLE_SUPER_ADMIN}:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="프로젝트 접근권한은 대표 또는 슈퍼관리자만 수정할 수 있어요",
        )


async def _seed_worker_documents(db: AsyncSession, worker_user_id: int) -> list[WorkerDocument]:
    result = await db.execute(
        select(WorkerDocument).where(WorkerDocument.worker_id == worker_user_id)
    )
    docs = list(result.scalars().all())
    if docs:
        return docs

    defaults = [
        ("doc_1", _WORKER_DOCUMENT_TYPE_ID_CARD, "신분증"),
        ("doc_2", _WORKER_DOCUMENT_TYPE_SAFETY_CERT, "안전교육 이수증"),
    ]
    for doc_id, document_type, name in defaults:
        db.add(
            WorkerDocument(
                worker_id=worker_user_id,
                document_id=doc_id,
                document_type=document_type,
                name=name,
                status="pending",
                review_status="pending_review",
            )
        )
    await db.flush()
    result = await db.execute(select(WorkerDocument).where(WorkerDocument.worker_id == worker_user_id))
    return list(result.scalars().all())


async def _seed_paystub_if_empty(db: AsyncSession, worker_user_id: int) -> None:
    count = (await db.execute(
        select(func.count()).select_from(Paystub).where(Paystub.worker_id == worker_user_id)
    )).scalar() or 0
    if count > 0:
        return

    month = datetime.utcnow().strftime("%Y-%m")
    paystub = Paystub(
        worker_id=worker_user_id,
        month=month,
        title=f"{month} 급여명세서",
        total_amount=2600000,
        deductions=180000,
        net_amount=2420000,
        status=PaystubStatus.SENT,
        date=datetime.utcnow().strftime("%Y-%m-%d"),
    )
    db.add(paystub)
    await db.flush()

    for label, amount in [
        ("기본급", 2200000),
        ("연장수당", 400000),
        ("소득세", -120000),
        ("지방세", -12000),
        ("고용보험", -48000),
    ]:
        db.add(PaystubItem(paystub_id=paystub.id, label=label, amount=amount))
    await db.flush()


async def _ensure_labor_contract_stub(
    db: AsyncSession,
    contract_id: int,
    *,
    worker_phone: Optional[str] = None,
) -> LaborContract:
    labor_contract = (await db.execute(select(LaborContract).where(LaborContract.id == contract_id))).scalar_one_or_none()
    if labor_contract:
        if worker_phone and not labor_contract.worker_phone:
            labor_contract.worker_phone = worker_phone
        return labor_contract

    project = (await db.execute(select(Project).order_by(Project.created_at.asc()).limit(1))).scalar_one_or_none()
    if not project:
        org = (await db.execute(select(Organization).order_by(Organization.created_at.asc()).limit(1))).scalar_one_or_none()
        if not org:
            org = Organization(name="기본 조직", is_active=True)
            db.add(org)
            await db.flush()

        project = Project(
            name="기본 현장",
            address="미설정",
            organization_id=org.id,
            created_by=None,
        )
        db.add(project)
        await db.flush()

    labor_contract = LaborContract(
        id=contract_id,
        project_id=project.id,
        worker_name="현장근로자",
        worker_phone=worker_phone,
        work_date=date.today(),
        work_type="보통인부",
        daily_rate=Decimal("150000"),
        status=LaborContractStatus.SENT,
    )
    db.add(labor_contract)
    await db.flush()
    return labor_contract


def _calc_worker_deductions(
    total_labor_cost: int,
    rates: InsuranceRate,
) -> dict[str, int]:
    gross = Decimal(total_labor_cost)

    def _round_down(value: Decimal) -> int:
        return int(value.quantize(Decimal("1"), rounding=ROUND_DOWN))

    taxable_base = max(Decimal("0"), gross - rates.income_deduction)
    income_tax = _round_down(taxable_base * rates.simplified_tax_rate)
    resident_tax = _round_down(Decimal(income_tax) * rates.local_tax_rate)
    health = _round_down(gross * rates.health_insurance_rate)
    care = _round_down(Decimal(health) * rates.longterm_care_rate)

    if gross < rates.pension_lower_limit:
        pension_base = Decimal("0")
    else:
        pension_base = min(gross, rates.pension_upper_limit)
    pension = _round_down(pension_base * rates.national_pension_rate)

    employment = _round_down(gross * rates.employment_insurance_rate)
    total_deductions = income_tax + resident_tax + health + care + pension + employment
    net_pay = max(0, total_labor_cost - total_deductions)
    return {
        "income_tax": income_tax,
        "resident_tax": resident_tax,
        "health_insurance": health,
        "longterm_care": care,
        "national_pension": pension,
        "employment_insurance": employment,
        "total_deductions": total_deductions,
        "net_pay": net_pay,
    }


async def _get_or_create_insurance_rate(
    db: AsyncSession,
    organization_id: int,
    year: int,
) -> InsuranceRate:
    result = await db.execute(
        select(InsuranceRate).where(
            InsuranceRate.organization_id == organization_id,
            InsuranceRate.effective_year == year,
        )
    )
    rate = result.scalar_one_or_none()
    if rate:
        return rate

    previous_rate = (
        await db.execute(
            select(InsuranceRate)
            .where(
                InsuranceRate.organization_id == organization_id,
                InsuranceRate.effective_year < year,
            )
            .order_by(InsuranceRate.effective_year.desc())
            .limit(1)
        )
    ).scalar_one_or_none()

    if not previous_rate:
        previous_rate = (
            await db.execute(
                select(InsuranceRate)
                .where(InsuranceRate.organization_id == organization_id)
                .order_by(InsuranceRate.effective_year.desc())
                .limit(1)
            )
        ).scalar_one_or_none()

    if previous_rate:
        rate = InsuranceRate(
            organization_id=organization_id,
            effective_year=year,
            income_deduction=previous_rate.income_deduction,
            simplified_tax_rate=previous_rate.simplified_tax_rate,
            local_tax_rate=previous_rate.local_tax_rate,
            employment_insurance_rate=previous_rate.employment_insurance_rate,
            health_insurance_rate=previous_rate.health_insurance_rate,
            longterm_care_rate=previous_rate.longterm_care_rate,
            national_pension_rate=previous_rate.national_pension_rate,
            pension_upper_limit=previous_rate.pension_upper_limit,
            pension_lower_limit=previous_rate.pension_lower_limit,
            health_premium_upper=previous_rate.health_premium_upper,
            health_premium_lower=previous_rate.health_premium_lower,
        )
    else:
        rate = InsuranceRate(organization_id=organization_id, effective_year=year)

    db.add(rate)
    await db.flush()
    return rate


# ----------------------------
# Project access / reports
# ----------------------------


class ProjectAccessUpdateRequest(BaseModel):
    manager_ids: list[str]


@router.get("/projects/{project_id}/access", response_model=APIResponse[dict])
async def get_project_access_policy(
    project_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    _require_project_policy_admin(current_user)
    await _get_project_for_user(db, project_id, current_user)

    result = await db.execute(
        select(ProjectAccessPolicy).where(ProjectAccessPolicy.project_id == project_id)
    )
    policy = result.scalar_one_or_none()

    manager_ids = [] if not policy else [str(mid) for mid in (policy.manager_ids or [])]
    return APIResponse.ok({"project_id": str(project_id), "manager_ids": manager_ids})


@router.put("/projects/{project_id}/access", response_model=APIResponse[dict])
async def update_project_access_policy(
    project_id: int,
    payload: ProjectAccessUpdateRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    _require_project_policy_admin(current_user)
    project = await _get_project_for_user(db, project_id, current_user)
    manager_ids = list(dict.fromkeys(_to_int(v) for v in payload.manager_ids))

    if manager_ids:
        managers = (
            await db.execute(
                select(User).where(
                    User.id.in_(manager_ids),
                    User.role == UserRole.SITE_MANAGER,
                )
            )
        ).scalars().all()
        valid_manager_ids = {
            int(manager.id)
            for manager in managers
            if manager.organization_id == project.organization_id
        }
        invalid_ids = [mid for mid in manager_ids if mid not in valid_manager_ids]
        if invalid_ids:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="같은 조직의 현장소장만 접근권한에 추가할 수 있어요",
            )

    result = await db.execute(
        select(ProjectAccessPolicy).where(ProjectAccessPolicy.project_id == project_id)
    )
    policy = result.scalar_one_or_none()
    now = datetime.utcnow()

    if not policy:
        policy = ProjectAccessPolicy(project_id=project_id, manager_ids=manager_ids, updated_at=now)
        db.add(policy)
    else:
        policy.manager_ids = manager_ids
        policy.updated_at = now

    await _log_activity(db, current_user.id, "settings_change", f"프로젝트 {project_id} 접근권한 수정")

    return APIResponse.ok({
        "project_id": str(project_id),
        "manager_ids": [str(mid) for mid in manager_ids],
    })


class DailyReportCreateRequest(BaseModel):
    work_date: date
    weather: Optional[str] = None
    temperature: Optional[str] = None
    work_description: str
    tomorrow_plan: Optional[str] = None
    photos: list[str] = []


@router.get("/projects/{project_id}/daily-reports", response_model=APIResponse[list[dict]])
async def list_daily_reports(
    project_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    await _get_project_for_user(db, project_id, current_user)

    result = await db.execute(
        select(DailyReport)
        .where(DailyReport.project_id == project_id)
        .order_by(DailyReport.work_date.desc(), DailyReport.created_at.desc())
    )
    reports = result.scalars().all()
    return APIResponse.ok([_serialize_daily_report(report) for report in reports])


@router.post("/projects/{project_id}/daily-reports", response_model=APIResponse[dict], status_code=status.HTTP_201_CREATED)
async def create_daily_report(
    project_id: int,
    payload: DailyReportCreateRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    await _get_project_for_user(db, project_id, current_user)

    report = DailyReport(
        project_id=project_id,
        work_date=payload.work_date,
        weather=payload.weather,
        temperature=payload.temperature,
        work_description=payload.work_description,
        tomorrow_plan=payload.tomorrow_plan,
        photos=payload.photos,
        created_by=current_user.id,
    )
    db.add(report)

    await _log_activity(db, current_user.id, "project_update", f"프로젝트 {project_id} 작업일지 등록")

    await db.flush()
    return APIResponse.ok(_serialize_daily_report(report))


@router.get("/projects/{project_id}/daily-reports/{report_id}", response_model=APIResponse[dict])
async def get_daily_report(
    project_id: int,
    report_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    await _get_project_for_user(db, project_id, current_user)
    report = await _get_daily_report_for_project(db, project_id, report_id)
    return APIResponse.ok(_serialize_daily_report(report))


@router.post("/projects/{project_id}/daily-reports/{report_id}/hwpx")
async def download_daily_report_hwpx(
    project_id: int,
    report_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    project = await _get_project_for_user(db, project_id, current_user)
    report = await _get_daily_report_for_project(db, project_id, report_id)

    template_path = _resolve_daily_report_template_path()
    if template_path is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="공사일지 HWPX 템플릿을 찾을 수 없어요",
        )

    context = _build_daily_report_hwpx_context(project, report)

    try:
        from app.services.hwpx_template_engine import HwpxTemplateEngine

        engine = HwpxTemplateEngine(strict=False)
        with tempfile.NamedTemporaryFile(suffix=".hwpx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            engine.render(
                template_path=template_path,
                output_path=tmp_path,
                context=context,
            )
            output_bytes = tmp_path.read_bytes()
        finally:
            tmp_path.unlink(missing_ok=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"작업일지 HWPX 생성에 실패했어요: {str(exc)}",
        ) from exc

    project_name = _sanitize_filename(project.name or "project")
    date_token = report.work_date.strftime("%Y%m%d")
    filename = f"공사일지_{project_name}_{date_token}.hwpx"
    encoded_filename = quote(filename, safe="")

    return StreamingResponse(
        io.BytesIO(output_bytes),
        media_type="application/vnd.hancom.hwpx",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
        },
    )


class UtilityStatusPatchRequest(BaseModel):
    status: Optional[UtilityStatus] = None
    doc_status: Optional[UtilityDocStatus] = None


@router.get("/projects/{project_id}/utilities", response_model=APIResponse[dict])
async def get_utilities(
    project_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    project = await _get_project_for_user(db, project_id, current_user)

    project_category = (project.category or "").strip().lower()
    if project_category != "school":
        return APIResponse.ok(
            {
                "enabled": False,
                "reason": "school_only",
                "message": "수도광열비는 학교 프로젝트에서만 관리해요.",
                "items": [],
                "timeline": [],
            }
        )

    if project.status not in {
        ProjectStatus.COMPLETED,
        ProjectStatus.WARRANTY,
        ProjectStatus.CLOSED,
    }:
        return APIResponse.ok(
            {
                "enabled": False,
                "reason": "completion_required",
                "message": "실준공/준공 이후에 수도광열비 정산을 시작해요.",
                "items": [],
                "timeline": [],
            }
        )

    await _seed_utilities_if_empty(db, project_id)

    items_result = await db.execute(
        select(UtilityItem)
        .where(UtilityItem.project_id == project_id)
        .order_by(UtilityItem.due_date.asc())
    )
    timeline_result = await db.execute(
        select(UtilityTimeline)
        .where(UtilityTimeline.project_id == project_id)
        .order_by(UtilityTimeline.date.desc())
    )

    items = items_result.scalars().all()
    timeline = timeline_result.scalars().all()

    return APIResponse.ok(
        {
            "enabled": True,
            "reason": None,
            "message": None,
            "items": [
                {
                    "id": str(item.id),
                    "type": item.type,
                    "month": item.month,
                    "status": item.status.value,
                    "amount": item.amount,
                    "due_date": item.due_date.isoformat(),
                    "doc_status": item.doc_status.value,
                }
                for item in items
            ],
            "timeline": [
                {
                    "id": str(t.id),
                    "date": t.date.isoformat(),
                    "message": t.message,
                }
                for t in timeline
            ],
        }
    )


@router.patch("/projects/{project_id}/utilities/{utility_id}", response_model=APIResponse[dict])
async def patch_utility_status(
    project_id: int,
    utility_id: int,
    payload: UtilityStatusPatchRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    project = await _get_project_for_user(db, project_id, current_user)

    project_category = (project.category or "").strip().lower()
    if project_category != "school":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="학교 프로젝트에서만 수도광열비 상태를 변경할 수 있어요.",
        )
    if project.status not in {
        ProjectStatus.COMPLETED,
        ProjectStatus.WARRANTY,
        ProjectStatus.CLOSED,
    }:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="준공 이후 프로젝트만 수도광열비 상태를 변경할 수 있어요.",
        )

    result = await db.execute(
        select(UtilityItem)
        .where(UtilityItem.id == utility_id)
        .where(UtilityItem.project_id == project_id)
    )
    item = result.scalar_one_or_none()
    if not item:
        raise NotFoundException("utility", utility_id)

    now = datetime.utcnow()
    if payload.status is not None:
        item.status = payload.status
    if payload.doc_status is not None:
        item.doc_status = payload.doc_status
    item.updated_at = now

    status_label = item.status.value
    db.add(
        UtilityTimeline(
            project_id=project_id,
            utility_item_id=item.id,
            date=now,
            message=f"{item.month} {item.type} 상태를 {status_label}로 변경했어요.",
        )
    )

    return APIResponse.ok({"id": str(item.id)})


class MaterialOrderItemInput(BaseModel):
    catalog_item_id: Optional[int] = None
    pricebook_revision_id: Optional[int] = None
    description: Optional[str] = None
    specification: Optional[str] = None
    unit: Optional[str] = None
    quantity: Decimal
    unit_price: Optional[Decimal] = None
    override_reason: Optional[str] = None


class MaterialOrderCreateRequest(BaseModel):
    items: list[MaterialOrderItemInput]
    notes: Optional[str] = None
    vendor_id: Optional[int] = None


class MaterialOrderStatusPatchRequest(BaseModel):
    status: str
    reason: Optional[str] = None
    invoice_number: Optional[str] = None
    invoice_amount: Optional[int] = None
    invoice_file_url: Optional[str] = None


@router.get("/projects/{project_id}/material-orders", response_model=APIResponse[list[dict]])
async def list_material_orders(
    project_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    await _get_project_for_user(db, project_id, current_user)

    result = await db.execute(
        select(MaterialOrder)
        .where(MaterialOrder.project_id == project_id)
        .order_by(MaterialOrder.created_at.desc())
    )
    orders = result.scalars().all()

    items_result = await db.execute(
        select(MaterialOrderItem)
        .where(MaterialOrderItem.material_order_id.in_([o.id for o in orders]))
    ) if orders else None
    item_rows = items_result.scalars().all() if items_result else []

    by_order: dict[int, list[MaterialOrderItem]] = {}
    for row in item_rows:
        by_order.setdefault(row.material_order_id, []).append(row)

    return APIResponse.ok(
        [_serialize_material_order(order, by_order.get(order.id, [])) for order in orders]
    )


@router.get("/projects/{project_id}/material-orders/mobile", response_model=APIResponse[list[dict]])
async def list_material_orders_mobile(
    project_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    await _get_project_for_user(db, project_id, current_user)

    result = await db.execute(
        select(MaterialOrder)
        .where(MaterialOrder.project_id == project_id)
        .order_by(MaterialOrder.updated_at.desc(), MaterialOrder.created_at.desc())
    )
    orders = result.scalars().all()
    if not orders:
        return APIResponse.ok([])

    item_counts_result = await db.execute(
        select(
            MaterialOrderItem.material_order_id,
            func.count(MaterialOrderItem.id),
        )
        .where(MaterialOrderItem.material_order_id.in_([o.id for o in orders]))
        .group_by(MaterialOrderItem.material_order_id)
    )
    item_count_map = {
        int(row[0]): int(row[1])
        for row in item_counts_result.all()
    }

    role = _role_value(current_user)
    show_amount = role in {ROLE_COMPANY_ADMIN, ROLE_SUPER_ADMIN}
    return APIResponse.ok([
        {
            "id": str(order.id),
            "order_number": order.order_number,
            "status": _normalize_material_order_status_value(order.status),
            "item_count": item_count_map.get(order.id, 0),
            "summary_amount": order.total_amount if show_amount else None,
            "requested_at": _to_iso(order.requested_at),
            "delivered_at": _to_iso(order.delivered_at),
            "updated_at": _to_iso(order.updated_at),
        }
        for order in orders
    ])


@router.post("/projects/{project_id}/material-orders", response_model=APIResponse[dict], status_code=status.HTTP_201_CREATED)
async def create_material_order(
    project_id: int,
    payload: MaterialOrderCreateRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    project = await _get_project_for_user(db, project_id, current_user)

    if not payload.items:
        raise HTTPException(status_code=400, detail="발주 품목이 필요해요")

    order_number = f"MO-{datetime.utcnow().strftime('%Y%m%d')}-{secrets.token_hex(2).upper()}"
    order = MaterialOrder(
        project_id=project_id,
        order_number=order_number,
        status=MaterialOrderStatus.DRAFT,
        vendor_id=payload.vendor_id,
        notes=payload.notes,
        created_by=current_user.id,
    )
    db.add(order)
    await db.flush()

    role = _role_value(current_user)
    can_override_price = role in {ROLE_COMPANY_ADMIN, ROLE_SUPER_ADMIN}
    total = Decimal("0")
    for row in payload.items:
        if row.quantity <= 0:
            raise HTTPException(status_code=400, detail="수량은 0보다 커야 해요")

        resolved_description = (row.description or "").strip()
        resolved_specification = (row.specification or "").strip() or None
        resolved_unit = (row.unit or "").strip()
        resolved_unit_price: Optional[Decimal] = row.unit_price
        resolved_catalog_item_id = row.catalog_item_id
        resolved_revision_id = row.pricebook_revision_id
        price_source = "catalog_revision"
        override_reason = (row.override_reason or "").strip() or None

        if row.catalog_item_id is not None:
            if resolved_revision_id is None:
                resolved_revision_id = project.pricebook_revision_id
            if resolved_revision_id is None:
                raise HTTPException(
                    status_code=400,
                    detail="단가표 버전이 연결되지 않아 자재 단가를 확인할 수 없어요.",
                )

            joined = (
                await db.execute(
                    select(CatalogItemPrice, CatalogItem)
                    .join(CatalogItem, CatalogItem.id == CatalogItemPrice.catalog_item_id)
                    .where(CatalogItemPrice.catalog_item_id == row.catalog_item_id)
                    .where(CatalogItemPrice.pricebook_revision_id == resolved_revision_id)
                )
            ).first()
            if not joined:
                raise HTTPException(
                    status_code=400,
                    detail="요청한 단가표 버전에서 자재 단가를 찾을 수 없어요.",
                )

            catalog_price, catalog_item = joined
            resolved_description = resolved_description or catalog_item.name_ko
            resolved_specification = (
                resolved_specification
                if resolved_specification is not None
                else (catalog_item.specification or None)
            )
            resolved_unit = resolved_unit or catalog_item.base_unit
            resolved_unit_price = catalog_price.unit_price

            if row.unit_price is not None and row.unit_price != catalog_price.unit_price:
                if not can_override_price:
                    raise HTTPException(
                        status_code=403,
                        detail="단가 수동 변경은 대표자 또는 최고관리자만 할 수 있어요.",
                    )
                if not override_reason:
                    raise HTTPException(
                        status_code=400,
                        detail="단가를 수동 변경할 때는 사유를 입력해 주세요.",
                    )
                resolved_unit_price = row.unit_price
                price_source = "manual_override"
        else:
            if not can_override_price:
                raise HTTPException(
                    status_code=400,
                    detail="카탈로그 품목을 선택해 주세요.",
                )
            if (
                not resolved_description
                or not resolved_unit
                or resolved_unit_price is None
                or resolved_unit_price <= 0
            ):
                raise HTTPException(
                    status_code=400,
                    detail="수동 품목은 품명, 단위, 단가를 모두 입력해 주세요.",
                )
            if not override_reason:
                raise HTTPException(
                    status_code=400,
                    detail="수동 품목 추가 시 사유를 입력해 주세요.",
                )
            price_source = "manual_override"
            resolved_catalog_item_id = None
            resolved_revision_id = None

        if resolved_unit_price is None or resolved_unit_price <= 0:
            raise HTTPException(status_code=400, detail="유효한 단가를 찾을 수 없어요.")
        if not resolved_description:
            raise HTTPException(status_code=400, detail="품목명을 입력해 주세요.")
        if not resolved_unit:
            raise HTTPException(status_code=400, detail="단위를 입력해 주세요.")

        amount = row.quantity * resolved_unit_price
        total += amount
        db.add(
            MaterialOrderItem(
                material_order_id=order.id,
                catalog_item_id=resolved_catalog_item_id,
                pricebook_revision_id=resolved_revision_id,
                price_source=price_source,
                override_reason=override_reason,
                description=resolved_description,
                specification=resolved_specification,
                unit=resolved_unit,
                quantity=row.quantity,
                unit_price=resolved_unit_price,
                amount=amount,
            )
        )

    order.total_amount = int(total)

    return APIResponse.ok(
        {
            "id": str(order.id),
            "order_number": order.order_number,
            "status": _normalize_material_order_status_value(order.status),
            "total_amount": order.total_amount,
        }
    )


@router.get("/material-orders/{order_id}", response_model=APIResponse[dict])
async def get_material_order(
    order_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    result = await db.execute(select(MaterialOrder).where(MaterialOrder.id == order_id))
    order = result.scalar_one_or_none()
    if not order:
        raise NotFoundException("material_order", order_id)

    await _get_project_for_user(db, order.project_id, current_user)

    item_result = await db.execute(
        select(MaterialOrderItem)
        .where(MaterialOrderItem.material_order_id == order.id)
        .order_by(MaterialOrderItem.id.asc())
    )
    order_items = item_result.scalars().all()

    return APIResponse.ok(_serialize_material_order(order, order_items))


@router.patch("/material-orders/{order_id}", response_model=APIResponse[dict])
async def update_material_order_status(
    order_id: int,
    payload: MaterialOrderStatusPatchRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    result = await db.execute(select(MaterialOrder).where(MaterialOrder.id == order_id))
    order = result.scalar_one_or_none()
    if not order:
        raise NotFoundException("material_order", order_id)

    await _get_project_for_user(db, order.project_id, current_user)

    try:
        requested_status = MaterialOrderStatus(payload.status)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="유효하지 않은 발주 상태예요") from exc

    current_status = _normalize_material_order_status_value(order.status)
    next_status_value = _normalize_material_order_status_value(requested_status)
    if current_status == next_status_value:
        return APIResponse.ok(
            {
                "id": str(order.id),
                "status": next_status_value,
                "message": "이미 같은 상태예요.",
            }
        )

    _require_material_order_status_permission(
        user=current_user,
        target_status=next_status_value,
    )

    transition_map = _material_order_allowed_transitions()
    allowed_targets = transition_map.get(current_status, set())
    if next_status_value not in allowed_targets:
        raise HTTPException(
            status_code=400,
            detail=f"{current_status} 상태에서 {next_status_value} 상태로는 변경할 수 없어요.",
        )

    if next_status_value == MaterialOrderStatus.CANCELLED.value and not (payload.reason or "").strip():
        raise HTTPException(status_code=400, detail="발주 취소 시 사유를 입력해 주세요.")

    order.status = MaterialOrderStatus(next_status_value)
    now = datetime.utcnow()
    if payload.invoice_number is not None:
        order.invoice_number = payload.invoice_number
    if payload.invoice_amount is not None:
        order.invoice_amount = payload.invoice_amount
    if payload.invoice_file_url is not None:
        order.invoice_file_url = payload.invoice_file_url

    if next_status_value == MaterialOrderStatus.REQUESTED.value:
        order.requested_at = now
    elif next_status_value == MaterialOrderStatus.INVOICE_RECEIVED.value:
        order.confirmed_at = now
    elif next_status_value == MaterialOrderStatus.PAYMENT_COMPLETED.value:
        order.payment_at = now
    elif next_status_value == MaterialOrderStatus.SHIPPED.value:
        order.shipped_at = now
    elif next_status_value == MaterialOrderStatus.DELIVERED.value:
        order.delivered_at = now
        order.received_at = now
        order.received_by_user_id = current_user.id
    elif next_status_value == MaterialOrderStatus.CLOSED.value:
        order.closed_at = now
    elif next_status_value == MaterialOrderStatus.CANCELLED.value:
        reason = (payload.reason or "").strip()
        if reason:
            note_prefix = f"[취소사유] {reason}"
            order.notes = f"{note_prefix}\n{order.notes}" if order.notes else note_prefix
    order.updated_at = now

    return APIResponse.ok(
        {
            "id": str(order.id),
            "status": _normalize_material_order_status_value(order.status),
            "message": "발주 상태를 변경했어요.",
        }
    )


@router.delete("/material-orders/{order_id}", response_model=APIResponse[dict])
async def cancel_material_order(
    order_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    result = await db.execute(select(MaterialOrder).where(MaterialOrder.id == order_id))
    order = result.scalar_one_or_none()
    if not order:
        raise NotFoundException("material_order", order_id)

    await _get_project_for_user(db, order.project_id, current_user)

    _require_material_order_status_permission(
        user=current_user,
        target_status=MaterialOrderStatus.CANCELLED.value,
    )
    current_status = _normalize_material_order_status_value(order.status)
    allowed_targets = _material_order_allowed_transitions().get(current_status, set())
    if MaterialOrderStatus.CANCELLED.value not in allowed_targets:
        raise HTTPException(
            status_code=400,
            detail="현재 상태에서는 발주를 취소할 수 없어요.",
        )

    order.status = MaterialOrderStatus.CANCELLED
    order.updated_at = datetime.utcnow()
    return APIResponse.ok({"id": str(order.id), "message": "발주를 취소했어요."})


@router.get("/projects/{project_id}/diagnoses", response_model=APIResponse[list[dict]])
async def list_project_diagnoses(
    project_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    await _get_project_for_user(db, project_id, current_user)

    visit_ids_result = await db.execute(
        select(SiteVisit.id).where(SiteVisit.project_id == project_id)
    )
    visit_ids = [row[0] for row in visit_ids_result.all()]

    if not visit_ids:
        return APIResponse.ok([])

    result = await db.execute(
        select(AIDiagnosis)
        .where(AIDiagnosis.site_visit_id.in_(visit_ids))
        .order_by(AIDiagnosis.created_at.desc())
    )
    diagnoses = result.scalars().all()

    return APIResponse.ok([
        {
            "id": str(d.id),
            "site_visit_id": str(d.site_visit_id),
            "status": d.status.value,
            "leak_opinion_text": d.leak_opinion_text,
            "created_at": d.created_at.isoformat(),
            "model_name": d.model_name,
        }
        for d in diagnoses
    ])


# ----------------------------
# Partners
# ----------------------------


class PartnerCreateRequest(BaseModel):
    name: str
    representative_name: Optional[str] = None
    representative_phone: Optional[str] = None
    business_number: Optional[str] = None
    contact_name: Optional[str] = None
    contact_phone: Optional[str] = None
    license_type: Optional[str] = None
    is_women_owned: Optional[bool] = None
    # Legacy compatibility fields
    owner: Optional[str] = None
    biz_no: Optional[str] = None
    license: Optional[str] = None
    is_female_owned: Optional[bool] = None


class PartnerUpdateRequest(BaseModel):
    name: Optional[str] = None
    representative_name: Optional[str] = None
    representative_phone: Optional[str] = None
    business_number: Optional[str] = None
    contact_name: Optional[str] = None
    contact_phone: Optional[str] = None
    license_type: Optional[str] = None
    is_women_owned: Optional[bool] = None
    # Legacy compatibility fields
    owner: Optional[str] = None
    biz_no: Optional[str] = None
    license: Optional[str] = None
    is_female_owned: Optional[bool] = None


def _normalize_partner_text(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    normalized = " ".join(value.strip().split())
    return normalized or None


def _resolve_partner_text(primary: Optional[str], legacy: Optional[str]) -> Optional[str]:
    return _normalize_partner_text(primary) or _normalize_partner_text(legacy)


def _partner_to_response(partner: Partner) -> dict:
    representative_name = partner.representative_name or partner.owner or ""
    business_number = partner.business_number or partner.biz_no or ""
    license_type = partner.license_type or partner.license or ""
    is_women_owned = bool(partner.is_women_owned or partner.is_female_owned)

    return {
        "id": str(partner.id),
        "name": partner.name,
        "representative_name": representative_name,
        "representative_phone": partner.representative_phone or "",
        "business_number": business_number,
        "contact_name": partner.contact_name or "",
        "contact_phone": partner.contact_phone or "",
        "license_type": license_type,
        "is_women_owned": is_women_owned,
        # Legacy compatibility
        "owner": representative_name,
        "biz_no": business_number,
        "license": license_type,
        "is_female_owned": is_women_owned,
        "status": partner.status.value,
    }


@router.get("/partners", response_model=APIResponse[list[dict]])
async def list_partners(
    db: DBSession,
    current_user: CurrentUser,
    search: Optional[str] = None,
    status_filter: Optional[str] = Query(default=None, alias="status"),
):
    org_id = _require_org_id(current_user)

    query = select(Partner).where(Partner.organization_id == org_id)
    if search:
        like = f"%{search}%"
        query = query.where(
            (Partner.name.ilike(like)) |
            (Partner.representative_name.ilike(like)) |
            (Partner.owner.ilike(like)) |
            (Partner.business_number.ilike(like)) |
            (Partner.biz_no.ilike(like)) |
            (Partner.representative_phone.ilike(like)) |
            (Partner.contact_name.ilike(like)) |
            (Partner.contact_phone.ilike(like))
        )
    if status_filter:
        try:
            query = query.where(Partner.status == PartnerStatus(status_filter))
        except ValueError:
            pass

    query = query.order_by(Partner.created_at.desc())
    result = await db.execute(query)
    items = result.scalars().all()

    return APIResponse.ok([_partner_to_response(p) for p in items])


@router.post("/partners", response_model=APIResponse[dict], status_code=status.HTTP_201_CREATED)
async def create_partner(
    payload: PartnerCreateRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    org_id = _require_org_id(current_user)

    name = _normalize_partner_text(payload.name)
    if not name:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="업체명을 입력해 주세요.",
        )

    representative_name = _resolve_partner_text(
        payload.representative_name, payload.owner
    )
    if not representative_name:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="대표자명을 입력해 주세요.",
        )

    business_number = _resolve_partner_text(payload.business_number, payload.biz_no)
    if not business_number:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="사업자번호를 입력해 주세요.",
        )

    license_type = _resolve_partner_text(payload.license_type, payload.license)
    is_women_owned = payload.is_women_owned
    if is_women_owned is None:
        is_women_owned = payload.is_female_owned if payload.is_female_owned is not None else False

    partner = Partner(
        organization_id=org_id,
        name=name,
        representative_name=representative_name,
        representative_phone=_normalize_partner_text(payload.representative_phone),
        business_number=business_number,
        contact_name=_normalize_partner_text(payload.contact_name),
        contact_phone=_normalize_partner_text(payload.contact_phone),
        license_type=license_type,
        is_women_owned=bool(is_women_owned),
        owner=representative_name,
        biz_no=business_number,
        license=license_type,
        is_female_owned=bool(is_women_owned),
        status=PartnerStatus.ACTIVE,
    )
    db.add(partner)
    await db.flush()

    return APIResponse.ok(_partner_to_response(partner))


@router.patch("/partners/{partner_id}", response_model=APIResponse[dict])
async def update_partner(
    partner_id: int,
    payload: PartnerUpdateRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    org_id = _require_org_id(current_user)

    result = await db.execute(
        select(Partner)
        .where(Partner.id == partner_id)
        .where(Partner.organization_id == org_id)
    )
    partner = result.scalar_one_or_none()
    if not partner:
        raise NotFoundException("partner", partner_id)

    updates = payload.model_dump(exclude_unset=True)

    if "name" in updates:
        normalized_name = _normalize_partner_text(updates["name"])
        if not normalized_name:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="업체명을 입력해 주세요.",
            )
        partner.name = normalized_name

    if "representative_name" in updates or "owner" in updates:
        representative_name = _resolve_partner_text(
            updates.get("representative_name"), updates.get("owner")
        )
        if not representative_name:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="대표자명을 입력해 주세요.",
            )
        partner.representative_name = representative_name
        partner.owner = representative_name

    if "business_number" in updates or "biz_no" in updates:
        business_number = _resolve_partner_text(
            updates.get("business_number"), updates.get("biz_no")
        )
        if not business_number:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="사업자번호를 입력해 주세요.",
            )
        partner.business_number = business_number
        partner.biz_no = business_number

    if "representative_phone" in updates:
        partner.representative_phone = _normalize_partner_text(
            updates["representative_phone"]
        )
    if "contact_name" in updates:
        partner.contact_name = _normalize_partner_text(updates["contact_name"])
    if "contact_phone" in updates:
        partner.contact_phone = _normalize_partner_text(updates["contact_phone"])

    if "license_type" in updates or "license" in updates:
        license_type = _resolve_partner_text(
            updates.get("license_type"), updates.get("license")
        )
        partner.license_type = license_type
        partner.license = license_type

    is_women_owned_update = updates.get("is_women_owned")
    if is_women_owned_update is None and "is_female_owned" in updates:
        is_women_owned_update = updates.get("is_female_owned")
    if is_women_owned_update is not None:
        partner.is_women_owned = bool(is_women_owned_update)
        partner.is_female_owned = bool(is_women_owned_update)

    partner.updated_at = datetime.utcnow()

    return APIResponse.ok(_partner_to_response(partner))


@router.delete("/partners/{partner_id}", response_model=APIResponse[dict])
async def delete_partner(
    partner_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    org_id = _require_org_id(current_user)

    result = await db.execute(
        select(Partner)
        .where(Partner.id == partner_id)
        .where(Partner.organization_id == org_id)
    )
    partner = result.scalar_one_or_none()
    if not partner:
        raise NotFoundException("partner", partner_id)

    await db.delete(partner)
    return APIResponse.ok({"deleted": True})


@router.post("/partners/{partner_id}/toggle-status", response_model=APIResponse[dict])
async def toggle_partner_status(
    partner_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    org_id = _require_org_id(current_user)

    result = await db.execute(
        select(Partner)
        .where(Partner.id == partner_id)
        .where(Partner.organization_id == org_id)
    )
    partner = result.scalar_one_or_none()
    if not partner:
        raise NotFoundException("partner", partner_id)

    partner.status = (
        PartnerStatus.INACTIVE if partner.status == PartnerStatus.ACTIVE else PartnerStatus.ACTIVE
    )
    partner.updated_at = datetime.utcnow()

    return APIResponse.ok(
        {
            "id": str(partner.id),
            "status": partner.status.value,
            "message": "협력사 상태를 변경했어요.",
        }
    )


# ----------------------------
# Invitations
# ----------------------------


class InvitationCreateRequest(BaseModel):
    phone: str
    name: str
    role: str
    organization_id: Optional[str] = None


class InvitationAcceptRequest(BaseModel):
    password: str


@router.post("/invitations", response_model=APIResponse[dict], status_code=status.HTTP_201_CREATED)
async def create_invitation(
    payload: InvitationCreateRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    ensure_can_create_invitation(current_user, payload.role)

    role = _role_value(current_user)
    if role == ROLE_COMPANY_ADMIN:
        org_id = _require_org_id(current_user)
    elif role == ROLE_SUPER_ADMIN:
        if not payload.organization_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="슈퍼관리자 초대에는 organization_id가 필요해요",
            )
        try:
            org_id = _to_int(payload.organization_id)
        except ValueError as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="organization_id 형식이 올바르지 않아요",
            ) from exc
        org = (await db.execute(select(Organization).where(Organization.id == org_id))).scalar_one_or_none()
        if not org:
            raise NotFoundException("organization", org_id)
    else:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="초대 권한이 없어요")

    token = secrets.token_urlsafe(24)
    invitation = Invitation(
        organization_id=org_id,
        phone=payload.phone,
        name=payload.name,
        role=payload.role,
        token=token,
        status=InvitationStatus.PENDING,
        created_by=current_user.id,
        expires_at=datetime.utcnow() + timedelta(days=7),
    )
    db.add(invitation)
    await db.flush()
    await _log_activity(
        db,
        current_user.id,
        "invitation_create",
        f"초대 생성: invitation_id={invitation.id}",
    )

    # 초대 알림톡 발송
    try:
        from app.services.sms import get_sms_service
        sms_service = get_sms_service()
        invite_url = f"/accept-invite/{token}"
        await sms_service.send_alimtalk(
            phone=payload.phone,
            template_code="USER_INVITE",
            variables={
                "name": payload.name or "고객",
                "invite_url": invite_url,
            },
        )
        await _log_activity(
            db,
            current_user.id,
            "notification_send_success",
            f"초대 알림톡 발송 성공: invitation_id={invitation.id}",
        )
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning(f"초대 알림톡 발송 실패 (무시): {e}")
        await _log_activity(
            db,
            current_user.id,
            "notification_send_failure",
            f"초대 알림톡 발송 실패: invitation_id={invitation.id}",
        )
    # 알림톡 실패해도 초대 생성은 성공으로 처리

    return APIResponse.ok(
        {
            "id": str(invitation.id),
            "token": invitation.token,
            "invite_url": f"/accept-invite/{invitation.token}",
        }
    )


@router.get("/invitations", response_model=PaginatedResponse[dict])
async def list_invitations(
    db: DBSession,
    current_user: CurrentUser,
    status_filter: Optional[InvitationStatus] = Query(default=None, alias="status"),
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=20, ge=1, le=100),
):
    org_id = _require_org_id(current_user)

    query = select(Invitation).where(Invitation.organization_id == org_id)
    if status_filter:
        query = query.where(Invitation.status == status_filter)

    total = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar() or 0

    result = await db.execute(
        query.order_by(Invitation.created_at.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
    )
    invitations = result.scalars().all()

    return PaginatedResponse.create(
        items=[
            {
                "id": str(i.id),
                "phone": i.phone,
                "name": i.name,
                "role": i.role,
                "status": i.status.value,
                "created_at": i.created_at.isoformat(),
                "expires_at": i.expires_at.isoformat(),
            }
            for i in invitations
        ],
        page=page,
        per_page=per_page,
        total=total,
    )


@router.post("/invitations/{invitation_id}/resend", response_model=APIResponse[dict])
async def resend_invitation(
    invitation_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    org_id = _require_org_id(current_user)

    result = await db.execute(
        select(Invitation)
        .where(Invitation.id == invitation_id)
        .where(Invitation.organization_id == org_id)
    )
    invitation = result.scalar_one_or_none()
    if not invitation:
        raise NotFoundException("invitation", invitation_id)

    invitation.token = secrets.token_urlsafe(24)
    invitation.expires_at = datetime.utcnow() + timedelta(days=7)
    invitation.status = InvitationStatus.PENDING
    await _log_activity(
        db,
        current_user.id,
        "invitation_resend",
        f"초대 재발송: invitation_id={invitation.id}",
    )

    # 재초대 알림톡 발송
    try:
        from app.services.sms import get_sms_service
        sms_service = get_sms_service()
        invite_url = f"/accept-invite/{invitation.token}"
        await sms_service.send_alimtalk(
            phone=invitation.phone,
            template_code="USER_INVITE",
            variables={
                "name": invitation.name or "고객",
                "invite_url": invite_url,
            },
        )
        await _log_activity(
            db,
            current_user.id,
            "notification_send_success",
            f"재초대 알림톡 발송 성공: invitation_id={invitation.id}",
        )
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning(f"재초대 알림톡 발송 실패 (무시): {e}")
        await _log_activity(
            db,
            current_user.id,
            "notification_send_failure",
            f"재초대 알림톡 발송 실패: invitation_id={invitation.id}",
        )
    # 알림톡 실패해도 재초대 처리는 성공으로 처리

    return APIResponse.ok(
        {
            "id": str(invitation.id),
            "token": invitation.token,
            "invite_url": f"/accept-invite/{invitation.token}",
        }
    )


@router.post("/invitations/{invitation_id}/revoke", response_model=APIResponse[dict])
async def revoke_invitation(
    invitation_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    org_id = _require_org_id(current_user)

    result = await db.execute(
        select(Invitation)
        .where(Invitation.id == invitation_id)
        .where(Invitation.organization_id == org_id)
    )
    invitation = result.scalar_one_or_none()
    if not invitation:
        raise NotFoundException("invitation", invitation_id)

    invitation.status = InvitationStatus.REVOKED
    await _log_activity(
        db,
        current_user.id,
        "invitation_revoke",
        f"초대 취소: invitation_id={invitation.id}",
    )

    return APIResponse.ok({"id": str(invitation.id), "status": "revoked"})


@router.get("/invitations/{token}", response_model=APIResponse[dict])
async def get_invitation_by_token(
    token: str,
    db: DBSession,
):
    result = await db.execute(select(Invitation).where(Invitation.token == token))
    invitation = result.scalar_one_or_none()
    if not invitation:
        raise NotFoundException("invitation", token)

    if invitation.status == InvitationStatus.REVOKED or invitation.expires_at < datetime.utcnow():
        return APIResponse.fail("INVITATION_EXPIRED", "만료되었거나 취소된 초대예요.")

    org = (await db.execute(select(Organization).where(Organization.id == invitation.organization_id))).scalar_one_or_none()

    return APIResponse.ok(
        {
            "id": str(invitation.id),
            "email": "",
            "name": invitation.name,
            "role": invitation.role,
            "organization_name": org.name if org else "",
            "status": invitation.status.value,
            "expires_at": invitation.expires_at.isoformat(),
        }
    )


@router.post("/invitations/{token}/accept", response_model=APIResponse[dict])
async def accept_invitation(
    token: str,
    payload: InvitationAcceptRequest,
    db: DBSession,
):
    result = await db.execute(select(Invitation).where(Invitation.token == token))
    invitation = result.scalar_one_or_none()
    if not invitation:
        raise NotFoundException("invitation", token)

    if invitation.status in {InvitationStatus.REVOKED, InvitationStatus.ACCEPTED}:
        raise HTTPException(status_code=400, detail="이미 처리된 초대예요.")
    if invitation.expires_at < datetime.utcnow():
        raise HTTPException(status_code=400, detail="만료된 초대예요.")

    existing = await db.execute(
        select(User).where(
            User.phone == invitation.phone,
            User.organization_id == invitation.organization_id,
        )
    )
    user = existing.scalar_one_or_none()

    if not user:
        username = f"{invitation.phone.replace('-', '')}_{secrets.token_hex(2)}"
        try:
            role = UserRole(invitation.role)
        except ValueError:
            role = UserRole.SITE_MANAGER

        user = User(
            username=username,
            email=None,
            name=invitation.name,
            phone=invitation.phone,
            role=role,
            organization_id=invitation.organization_id,
            password_hash=get_password_hash(payload.password),
            is_active=True,
        )
        db.add(user)
        await db.flush()
    else:
        user.password_hash = get_password_hash(payload.password)
        user.is_active = True

    invitation.status = InvitationStatus.ACCEPTED
    invitation.accepted_at = datetime.utcnow()
    invitation.accepted_user_id = user.id
    await write_activity_log(
        db,
        user_id=user.id,
        action="invitation_accept",
        description=f"초대 수락: invitation_id={invitation.id}",
    )

    return APIResponse.ok({"id": str(invitation.id), "status": invitation.status.value})


# ----------------------------
# My page
# ----------------------------


class ProfileUpdateRequest(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None


class NotificationPrefsUpdateRequest(BaseModel):
    email_notifications: bool
    project_status_change: bool
    estimate_contract_alerts: bool
    daily_report_alerts: bool
    platform_announcements: bool


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


class AccountDeletionRequest(BaseModel):
    password: str
    reason: Optional[str] = None


@router.patch("/me/profile", response_model=APIResponse[dict])
async def update_my_profile(
    payload: ProfileUpdateRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    updates = payload.model_dump(exclude_unset=True)
    for field, value in updates.items():
        setattr(current_user, field, value)
    current_user.updated_at = datetime.utcnow()
    await _log_activity(db, current_user.id, "profile_update", "프로필 수정")

    return APIResponse.ok(
        {
            "id": str(current_user.id),
            "name": current_user.name,
            "email": current_user.email,
            "phone": current_user.phone,
        }
    )


@router.get("/me/notification-prefs", response_model=APIResponse[dict])
async def get_my_notification_prefs(
    db: DBSession,
    current_user: CurrentUser,
):
    prefs = await _get_or_create_notification_prefs(db, current_user.id)
    return APIResponse.ok(
        {
            "user_id": str(current_user.id),
            "email_notifications": prefs.email_notifications,
            "project_status_change": prefs.project_status_change,
            "estimate_contract_alerts": prefs.estimate_contract_alerts,
            "daily_report_alerts": prefs.daily_report_alerts,
            "platform_announcements": prefs.platform_announcements,
        }
    )


@router.put("/me/notification-prefs", response_model=APIResponse[dict])
async def update_my_notification_prefs(
    payload: NotificationPrefsUpdateRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    prefs = await _get_or_create_notification_prefs(db, current_user.id)

    prefs.email_notifications = payload.email_notifications
    prefs.project_status_change = payload.project_status_change
    prefs.estimate_contract_alerts = payload.estimate_contract_alerts
    prefs.daily_report_alerts = payload.daily_report_alerts
    prefs.platform_announcements = payload.platform_announcements
    prefs.updated_at = datetime.utcnow()

    await _log_activity(db, current_user.id, "settings_change", "알림 설정 변경")

    return APIResponse.ok(
        {
            "user_id": str(current_user.id),
            "email_notifications": prefs.email_notifications,
            "project_status_change": prefs.project_status_change,
            "estimate_contract_alerts": prefs.estimate_contract_alerts,
            "daily_report_alerts": prefs.daily_report_alerts,
            "platform_announcements": prefs.platform_announcements,
        }
    )


@router.get("/me/activity-log", response_model=PaginatedResponse[dict])
async def get_my_activity_log(
    db: DBSession,
    current_user: CurrentUser,
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=20, ge=1, le=100),
):
    query = select(ActivityLog).where(ActivityLog.user_id == current_user.id)
    total = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar() or 0

    result = await db.execute(
        query.order_by(ActivityLog.created_at.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
    )
    logs = result.scalars().all()

    return PaginatedResponse.create(
        items=[
            {
                "id": str(log.id),
                "user_id": str(log.user_id),
                "action": log.action,
                "description": log.description,
                "ip_address": log.ip_address,
                "device_info": log.device_info,
                "created_at": log.created_at.isoformat(),
            }
            for log in logs
        ],
        page=page,
        per_page=per_page,
        total=total,
    )


@router.post("/me/change-password", response_model=APIResponse[dict])
async def change_my_password(
    payload: ChangePasswordRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    if not verify_password(payload.current_password, current_user.password_hash):
        raise HTTPException(status_code=400, detail="현재 비밀번호가 올바르지 않아요")

    current_user.password_hash = get_password_hash(payload.new_password)
    current_user.updated_at = datetime.utcnow()

    await _log_activity(db, current_user.id, "password_change", "비밀번호 변경")

    return APIResponse.ok({"changed": True})


@router.post("/me/logout-all-devices", response_model=APIResponse[dict])
async def logout_all_devices(
    db: DBSession,
    current_user: CurrentUser,
):
    await _log_activity(db, current_user.id, "logout", "모든 기기 로그아웃")
    return APIResponse.ok({"logged_out": True})


@router.post("/me/account-deactivation", response_model=APIResponse[dict])
async def request_account_deactivation(
    db: DBSession,
    current_user: CurrentUser,
):
    current_user.is_active = False
    current_user.updated_at = datetime.utcnow()
    db.add(AccountRequest(user_id=current_user.id, type="deactivation", status="requested"))
    await _log_activity(db, current_user.id, "settings_change", "계정 비활성화")

    return APIResponse.ok({"requested": True})


@router.post("/me/account-deletion", response_model=APIResponse[dict])
async def request_account_deletion(
    payload: AccountDeletionRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    if not verify_password(payload.password, current_user.password_hash):
        raise HTTPException(status_code=400, detail="비밀번호가 올바르지 않아요")

    current_user.is_active = False
    current_user.updated_at = datetime.utcnow()

    db.add(
        AccountRequest(
            user_id=current_user.id,
            type="deletion",
            reason=payload.reason,
            status="requested",
        )
    )
    await _log_activity(db, current_user.id, "settings_change", "회원탈퇴 요청")

    return APIResponse.ok({"requested": True})


# ----------------------------
# Labor domain
# ----------------------------


@router.get("/labor/overview", response_model=APIResponse[dict])
async def get_labor_overview(
    db: DBSession,
    current_user: CurrentUser,
):
    role = _role_value(current_user)

    worker_query = select(DailyWorker)
    if current_user.organization_id is not None:
        worker_query = worker_query.where(DailyWorker.organization_id == current_user.organization_id)

    workers = (await db.execute(worker_query.order_by(DailyWorker.created_at.desc()))).scalars().all()

    unsigned_contracts_query = select(func.count()).select_from(LaborContract).where(
        LaborContract.status.in_([LaborContractStatus.DRAFT, LaborContractStatus.SENT])
    )

    pending_paystubs_query = select(func.count()).select_from(Paystub).where(Paystub.status == PaystubStatus.SENT)

    if current_user.organization_id is not None:
        project_ids = [row[0] for row in (await db.execute(select(Project.id).where(Project.organization_id == current_user.organization_id))).all()]
        if project_ids:
            unsigned_contracts_query = unsigned_contracts_query.where(LaborContract.project_id.in_(project_ids))
        else:
            unsigned_contracts_query = unsigned_contracts_query.where(False)

    unsigned_contracts = (await db.execute(unsigned_contracts_query)).scalar() or 0
    pending_paystubs = (await db.execute(pending_paystubs_query)).scalar() or 0

    today = date.today()
    worker_rows = []
    for worker in workers:
        worker_rows.append(
            {
                "id": str(worker.id),
                "name": worker.name,
                "role": worker.job_type,
                "status": "inactive" if worker.is_blocked_for_labor else "active",
                "contract_status": "signed" if worker.registration_status == "registered" else "pending",
                "last_work_date": today.isoformat(),
                "is_blocked_for_labor": worker.is_blocked_for_labor,
                "block_reason": worker.block_reason,
            }
        )

    return APIResponse.ok(
        {
            "summary": {
                "active_workers": len(workers),
                "pending_paystubs": pending_paystubs,
                "unsigned_contracts": unsigned_contracts,
            },
            "workers": worker_rows,
        }
    )


@router.get("/admin/labor/overview", response_model=APIResponse[dict])
async def get_admin_labor_overview(
    db: DBSession,
    current_user: CurrentUser,
):
    if _role_value(current_user) != "super_admin":
        raise HTTPException(status_code=403, detail="슈퍼관리자만 접근 가능해요")

    workers_result = await db.execute(
        select(DailyWorker, Organization)
        .outerjoin(Organization, Organization.id == DailyWorker.organization_id)
        .order_by(DailyWorker.created_at.desc())
        .limit(200)
    )
    workers_joined = workers_result.all()

    active_workers = (
        await db.execute(select(func.count()).select_from(DailyWorker))
    ).scalar() or 0
    pending_paystubs = (
        await db.execute(
            select(func.count())
            .select_from(Paystub)
            .where(Paystub.status == PaystubStatus.SENT)
        )
    ).scalar() or 0
    unsigned_contracts = (
        await db.execute(
            select(func.count())
            .select_from(LaborContract)
            .where(
                LaborContract.status.in_(
                    [LaborContractStatus.DRAFT, LaborContractStatus.SENT]
                )
            )
        )
    ).scalar() or 0

    tenant_distribution_raw = (
        await db.execute(
            select(
                Organization.id,
                Organization.name,
                func.count(DailyWorker.id),
            )
            .join(DailyWorker, DailyWorker.organization_id == Organization.id)
            .group_by(Organization.id, Organization.name)
            .order_by(func.count(DailyWorker.id).desc())
            .limit(10)
        )
    ).all()

    today = date.today().isoformat()
    workers = [
        {
            "id": str(worker.id),
            "name": worker.name,
            "role": worker.job_type,
            "organization_id": str(worker.organization_id),
            "organization_name": org.name if org else "미지정",
            "status": "inactive" if worker.is_blocked_for_labor else "active",
            "contract_status": "signed" if worker.registration_status == "registered" else "pending",
            "last_work_date": today,
            "is_blocked_for_labor": worker.is_blocked_for_labor,
            "block_reason": worker.block_reason,
        }
        for worker, org in workers_joined
    ]

    tenant_worker_distribution = [
        {
            "organization_id": str(org_id),
            "organization_name": org_name,
            "worker_count": int(worker_count or 0),
        }
        for org_id, org_name, worker_count in tenant_distribution_raw
    ]

    return APIResponse.ok(
        {
            "summary": {
                "active_workers": int(active_workers),
                "pending_paystubs": int(pending_paystubs),
                "unsigned_contracts": int(unsigned_contracts),
                "organizations_with_workers": len(tenant_worker_distribution),
            },
            "workers": workers,
            "tenant_worker_distribution": tenant_worker_distribution,
        }
    )


@router.post("/labor/paystubs/batch-send", response_model=APIResponse[dict])
async def batch_send_paystubs(
    db: DBSession,
    current_user: CurrentUser,
):
    _require_non_super_admin_for_labor_actions(current_user)

    query = select(Paystub).where(Paystub.status == PaystubStatus.SENT)
    if current_user.organization_id is not None:
        query = query.join(DailyWorker, DailyWorker.id == Paystub.worker_id).where(
            DailyWorker.organization_id == current_user.organization_id
        )
    pending = (await db.execute(query)).scalars().all()

    for paystub in pending:
        paystub.status = PaystubStatus.CONFIRMED

    return APIResponse.ok(
        {
            "sent_count": len(pending),
            "message": f"{len(pending)}건의 지급명세서를 발송했어요.",
        }
    )


@router.get("/labor/codebook", response_model=APIResponse[dict])
async def get_labor_codebook(
    current_user: CurrentUser,
):
    return APIResponse.ok(get_labor_codebook_payload())


class DailyWorkerUpsertRequest(BaseModel):
    name: str
    job_type: str = "보통인부"
    job_type_code: str = ""
    team: str = ""
    hire_date: str
    visa_status: Optional[str] = None
    nationality_code: Optional[str] = None
    english_name: Optional[str] = None
    birth_date: str
    gender: int
    address: str = ""
    daily_rate: Decimal
    account_number: str = ""
    bank_name: str = ""
    phone: str = ""
    is_foreign: bool = False
    has_id_card: Optional[bool] = None
    has_safety_cert: Optional[bool] = None
    organization_id: Optional[str] = None


@router.get("/labor/daily-workers", response_model=APIResponse[list[dict]])
async def list_daily_workers(
    db: DBSession,
    current_user: CurrentUser,
):
    query = select(DailyWorker)
    if current_user.organization_id is not None:
        query = query.where(DailyWorker.organization_id == current_user.organization_id)
    query = query.order_by(DailyWorker.created_at.desc())

    workers = (await db.execute(query)).scalars().all()

    return APIResponse.ok([
        {
            "id": str(w.id),
            "name": w.name,
            "job_type": w.job_type,
            "job_type_code": w.job_type_code,
            "team": w.team,
            "hire_date": w.hire_date.isoformat(),
            "visa_status": w.visa_status,
            "nationality_code": w.nationality_code,
            "english_name": w.english_name,
            "birth_date": w.birth_date,
            "gender": w.gender,
            "address": w.address,
            "daily_rate": int(w.daily_rate),
            "account_number": w.account_number,
            "bank_name": w.bank_name,
            "phone": w.phone,
            "is_foreign": w.is_foreign,
            "organization_id": str(w.organization_id),
            "registration_status": w.registration_status,
            "invite_token": w.invite_token,
            "has_id_card": w.has_id_card,
            "has_safety_cert": w.has_safety_cert,
            "is_blocked_for_labor": w.is_blocked_for_labor,
            "block_reason": w.block_reason,
            "blocked_by_user_id": str(w.blocked_by_user_id) if w.blocked_by_user_id else None,
            "blocked_at": _to_iso(w.blocked_at),
        }
        for w in workers
    ])


@router.post("/labor/daily-workers", response_model=APIResponse[dict])
async def create_daily_worker(
    payload: DailyWorkerUpsertRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    _require_non_super_admin_for_labor_actions(current_user)
    _validate_daily_worker_payload(payload)

    if current_user.organization_id is not None:
        org_id = current_user.organization_id
    else:
        if not payload.organization_id:
            raise HTTPException(status_code=400, detail="organization_id가 필요해요")
        try:
            org_id = _to_int(payload.organization_id)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="organization_id 형식이 올바르지 않아요") from exc

    worker = DailyWorker(
        organization_id=org_id,
        name=payload.name,
        job_type=payload.job_type,
        job_type_code=normalize_code(payload.job_type_code),
        team=payload.team,
        hire_date=date.fromisoformat(payload.hire_date),
        visa_status=_normalized_or_none(payload.visa_status) if payload.is_foreign else None,
        nationality_code=_normalized_or_none(payload.nationality_code) if payload.is_foreign else None,
        english_name=payload.english_name,
        birth_date=payload.birth_date,
        gender=payload.gender,
        address=payload.address,
        daily_rate=payload.daily_rate,
        account_number=payload.account_number,
        bank_name=payload.bank_name,
        phone=payload.phone,
        is_foreign=payload.is_foreign,
        has_id_card=bool(payload.has_id_card),
        has_safety_cert=bool(payload.has_safety_cert),
    )
    db.add(worker)
    await db.flush()

    return APIResponse.ok({"id": str(worker.id)})


@router.patch("/labor/daily-workers/{worker_id}", response_model=APIResponse[dict])
async def update_daily_worker(
    worker_id: int,
    payload: DailyWorkerUpsertRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    _require_non_super_admin_for_labor_actions(current_user)
    _validate_daily_worker_payload(payload)

    query = select(DailyWorker).where(DailyWorker.id == worker_id)
    if current_user.organization_id is not None:
        query = query.where(DailyWorker.organization_id == current_user.organization_id)

    worker = (await db.execute(query)).scalar_one_or_none()
    if not worker:
        raise NotFoundException("daily_worker", worker_id)

    worker.name = payload.name
    worker.job_type = payload.job_type
    worker.job_type_code = normalize_code(payload.job_type_code)
    worker.team = payload.team
    worker.hire_date = date.fromisoformat(payload.hire_date)
    worker.visa_status = _normalized_or_none(payload.visa_status) if payload.is_foreign else None
    worker.nationality_code = _normalized_or_none(payload.nationality_code) if payload.is_foreign else None
    worker.english_name = payload.english_name
    worker.birth_date = payload.birth_date
    worker.gender = payload.gender
    worker.address = payload.address
    worker.daily_rate = payload.daily_rate
    worker.account_number = payload.account_number
    worker.bank_name = payload.bank_name
    worker.phone = payload.phone
    worker.is_foreign = payload.is_foreign
    if payload.has_id_card is not None:
        worker.has_id_card = payload.has_id_card
    if payload.has_safety_cert is not None:
        worker.has_safety_cert = payload.has_safety_cert
    worker.updated_at = datetime.utcnow()

    return APIResponse.ok({"id": str(worker.id)})


@router.delete("/labor/daily-workers/{worker_id}", response_model=APIResponse[dict])
async def delete_daily_worker(
    worker_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    _require_non_super_admin_for_labor_actions(current_user)
    query = select(DailyWorker).where(DailyWorker.id == worker_id)
    if current_user.organization_id is not None:
        query = query.where(DailyWorker.organization_id == current_user.organization_id)

    worker = (await db.execute(query)).scalar_one_or_none()
    if not worker:
        raise NotFoundException("daily_worker", worker_id)

    await db.delete(worker)
    return APIResponse.ok({"id": str(worker.id)})


class WorkerDocumentReviewRequest(BaseModel):
    action: str
    reason: Optional[str] = None


class DailyWorkerControlRequest(BaseModel):
    action: str
    reason: Optional[str] = None


@router.get("/labor/daily-workers/{worker_id}/documents", response_model=APIResponse[list[dict]])
async def get_daily_worker_documents(
    worker_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    _require_worker_document_upload_role(current_user)
    worker = await _get_daily_worker_for_document_ops(db, worker_id, current_user)

    docs = (
        await db.execute(
            select(WorkerDocument).where(WorkerDocument.worker_id == worker.id)
        )
    ).scalars().all()

    by_type: dict[str, WorkerDocument] = {}
    for doc in docs:
        doc_type = _normalize_worker_document_type(doc.document_type or doc.document_id)
        if doc_type not in _WORKER_DOCUMENT_ALLOWED_TYPES:
            continue

        existing = by_type.get(doc_type)
        if not existing:
            by_type[doc_type] = doc
            continue

        existing_uploaded_at = existing.uploaded_at or datetime.min
        current_uploaded_at = doc.uploaded_at or datetime.min
        if current_uploaded_at >= existing_uploaded_at:
            by_type[doc_type] = doc

    payload: list[dict] = []
    for doc_type in sorted(_WORKER_DOCUMENT_ALLOWED_TYPES):
        doc = by_type.get(doc_type)
        if doc:
            payload.append(_serialize_worker_document(doc))
            continue
        payload.append(
            {
                "id": None,
                "worker_id": str(worker.id),
                "organization_id": str(worker.organization_id),
                "document_id": doc_type,
                "document_type": doc_type,
                "document_name": _worker_document_label(doc_type),
                "name": _worker_document_label(doc_type),
                "status": "pending",
                "storage_path": None,
                "original_filename": None,
                "mime_type": None,
                "file_size_bytes": 0,
                "file_hash_sha256": None,
                "review_status": "pending_review",
                "review_reason": None,
                "reviewed_by_user_id": None,
                "reviewed_at": None,
                "anomaly_flags": [],
                "uploaded_at": None,
            }
        )

    return APIResponse.ok(payload)


@router.post("/labor/daily-workers/{worker_id}/documents/{document_type}", response_model=APIResponse[dict])
async def upload_daily_worker_document(
    worker_id: int,
    document_type: str,
    db: DBSession,
    current_user: CurrentUser,
    file: UploadFile = File(...),
):
    _require_worker_document_upload_role(current_user)
    normalized_doc_type = _require_valid_worker_document_type(document_type)
    worker = await _get_daily_worker_for_document_ops(db, worker_id, current_user)
    ext, mime_type = _validate_worker_document_file(file)

    content = await file.read()
    await file.seek(0)
    if len(content) == 0:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="빈 파일은 업로드할 수 없어요.",
        )

    file_hash = hashlib.sha256(content).hexdigest()
    duplicate_hash_exists = (
        await db.execute(
            select(WorkerDocument.id)
            .where(WorkerDocument.file_hash_sha256 == file_hash)
            .where(WorkerDocument.worker_id != worker.id)
            .limit(1)
        )
    ).scalar_one_or_none() is not None
    anomaly_flags = _detect_worker_document_anomaly_flags(
        content=content,
        ext=ext,
        mime_type=mime_type,
        duplicate_hash_exists=duplicate_hash_exists,
    )

    storage_path = await storage_service.save_bytes(
        data=content,
        category="worker-documents",
        subfolder=f"org-{worker.organization_id}/workers/{worker.id}",
        filename=f"{normalized_doc_type}_{int(datetime.utcnow().timestamp())}{ext}",
    )

    worker_docs = (
        await db.execute(
            select(WorkerDocument).where(WorkerDocument.worker_id == worker.id)
        )
    ).scalars().all()
    target: Optional[WorkerDocument] = None
    for candidate in worker_docs:
        candidate_type = _normalize_worker_document_type(
            candidate.document_type or candidate.document_id
        )
        if candidate_type == normalized_doc_type:
            target = candidate
            break

    original_filename = file.filename or f"{normalized_doc_type}{ext}"
    now_utc = datetime.utcnow()
    if target is None:
        target = WorkerDocument(
            worker_id=worker.id,
            organization_id=worker.organization_id,
            document_id=normalized_doc_type,
            document_type=normalized_doc_type,
            name=_worker_document_label(normalized_doc_type),
            status="submitted",
            storage_path=storage_path,
            original_filename=original_filename,
            mime_type=mime_type or None,
            file_size_bytes=len(content),
            file_hash_sha256=file_hash,
            review_status="pending_review",
            review_reason=None,
            reviewed_by_user_id=None,
            reviewed_at=None,
            anomaly_flags=anomaly_flags,
            uploaded_at=now_utc,
        )
        db.add(target)
        await db.flush()
    else:
        target.organization_id = worker.organization_id
        target.document_type = normalized_doc_type
        target.status = "submitted"
        target.storage_path = storage_path
        target.original_filename = original_filename
        target.mime_type = mime_type or None
        target.file_size_bytes = len(content)
        target.file_hash_sha256 = file_hash
        target.review_status = "pending_review"
        target.review_reason = None
        target.reviewed_by_user_id = None
        target.reviewed_at = None
        target.anomaly_flags = anomaly_flags
        target.uploaded_at = now_utc
        await db.flush()

    await _sync_daily_worker_document_flags(db, worker)
    await _write_worker_document_moderation_log(
        db,
        organization_id=worker.organization_id,
        worker_id=worker.id,
        worker_document_id=target.id,
        actor_user_id=current_user.id,
        action="upload",
        reason=None,
        payload={
            "document_type": normalized_doc_type,
            "storage_path": storage_path,
            "anomaly_flags": anomaly_flags,
        },
    )
    await _log_activity(
        db,
        current_user.id,
        "worker_document_upload",
        f"근로자 서류 업로드: worker_id={worker.id}, type={normalized_doc_type}",
    )

    return APIResponse.ok(
        {
            "worker_id": str(worker.id),
            "document": _serialize_worker_document(target),
            "requires_review": True,
        }
    )


@router.get("/labor/worker-documents/review-queue", response_model=APIResponse[list[dict]])
async def get_worker_document_review_queue(
    db: DBSession,
    current_user: CurrentUser,
    review_status: Optional[str] = Query(default="pending_review"),
):
    _require_worker_document_review_role(current_user)

    normalized_review_status = (review_status or "pending_review").strip().lower()
    if normalized_review_status != "all" and normalized_review_status not in _WORKER_DOCUMENT_REVIEW_STATUSES:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="지원하지 않는 검토 상태 필터예요.",
        )

    query = (
        select(WorkerDocument, DailyWorker)
        .join(DailyWorker, DailyWorker.id == WorkerDocument.worker_id)
        .where(WorkerDocument.document_type.in_(tuple(sorted(_WORKER_DOCUMENT_ALLOWED_TYPES))))
    )
    if current_user.organization_id is not None:
        query = query.where(DailyWorker.organization_id == current_user.organization_id)
    if normalized_review_status != "all":
        query = query.where(WorkerDocument.review_status == normalized_review_status)

    rows = (
        await db.execute(
            query.order_by(WorkerDocument.uploaded_at.desc(), WorkerDocument.id.desc()).limit(300)
        )
    ).all()

    payload: list[dict] = []
    for document, worker in rows:
        serialized = _serialize_worker_document(document)
        serialized.update(
            {
                "worker_name": worker.name,
                "worker_job_type": worker.job_type,
                "worker_registration_status": worker.registration_status,
                "worker_is_blocked_for_labor": worker.is_blocked_for_labor,
                "worker_block_reason": worker.block_reason,
            }
        )
        payload.append(serialized)

    return APIResponse.ok(payload)


@router.post("/labor/worker-documents/{worker_document_id}/review", response_model=APIResponse[dict])
async def review_worker_document(
    worker_document_id: int,
    payload: WorkerDocumentReviewRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    _require_worker_document_review_role(current_user)

    action = (payload.action or "").strip().lower()
    if action not in _WORKER_DOCUMENT_REVIEW_ACTIONS:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="지원하지 않는 검토 액션이에요.",
        )

    reason = (payload.reason or "").strip() or None
    if action in {"reject", "quarantine", "request_reupload"} and not reason:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="반려/격리/재업로드 요청 시 사유를 입력해주세요.",
        )

    document = (
        await db.execute(select(WorkerDocument).where(WorkerDocument.id == worker_document_id))
    ).scalar_one_or_none()
    if not document:
        raise NotFoundException("worker_document", worker_document_id)

    normalized_doc_type = _normalize_worker_document_type(
        document.document_type or document.document_id
    )
    if normalized_doc_type not in _WORKER_DOCUMENT_ALLOWED_TYPES:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="검토 가능한 근로자 서류가 아니에요.",
        )

    worker = (
        await db.execute(select(DailyWorker).where(DailyWorker.id == document.worker_id))
    ).scalar_one_or_none()
    if not worker:
        raise NotFoundException("daily_worker", document.worker_id)
    if current_user.organization_id is not None and worker.organization_id != current_user.organization_id:
        raise NotFoundException("worker_document", worker_document_id)

    if document.organization_id is None:
        document.organization_id = worker.organization_id

    status_by_action = {
        "approve": "approved",
        "reject": "rejected",
        "quarantine": "quarantined",
        "request_reupload": "rejected",
    }
    document.review_status = status_by_action[action]
    document.review_reason = reason
    document.reviewed_by_user_id = current_user.id
    document.reviewed_at = datetime.utcnow()
    if action == "approve":
        document.status = "submitted"
    elif action == "quarantine":
        document.status = "blocked"
    else:
        document.status = "rejected"

    await _sync_daily_worker_document_flags(db, worker)
    await _write_worker_document_moderation_log(
        db,
        organization_id=worker.organization_id,
        worker_id=worker.id,
        worker_document_id=document.id,
        actor_user_id=current_user.id,
        action=action,
        reason=reason,
        payload={
            "review_status": document.review_status,
            "document_type": normalized_doc_type,
        },
    )
    await _log_activity(
        db,
        current_user.id,
        "worker_document_review",
        f"근로자 서류 검토: worker_id={worker.id}, action={action}",
    )

    return APIResponse.ok(
        {
            "worker_id": str(worker.id),
            "document": _serialize_worker_document(document),
        }
    )


@router.patch("/labor/daily-workers/{worker_id}/control", response_model=APIResponse[dict])
async def control_daily_worker(
    worker_id: int,
    payload: DailyWorkerControlRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    _require_worker_document_review_role(current_user)

    action = (payload.action or "").strip().lower()
    if action not in _WORKER_DOCUMENT_CONTROL_ACTIONS:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="지원하지 않는 근로자 통제 액션이에요.",
        )

    reason = (payload.reason or "").strip() or None
    if action == "block" and not reason:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="차단 사유를 입력해주세요.",
        )

    worker = await _get_daily_worker_for_document_ops(db, worker_id, current_user)
    now_utc = datetime.utcnow()
    if action == "block":
        worker.is_blocked_for_labor = True
        worker.block_reason = reason
        worker.blocked_by_user_id = current_user.id
        worker.blocked_at = now_utc
    else:
        worker.is_blocked_for_labor = False
        worker.block_reason = None
        worker.blocked_by_user_id = None
        worker.blocked_at = None
    worker.updated_at = now_utc

    await _write_worker_document_moderation_log(
        db,
        organization_id=worker.organization_id,
        worker_id=worker.id,
        worker_document_id=None,
        actor_user_id=current_user.id,
        action=f"worker_{action}",
        reason=reason,
        payload={
            "is_blocked_for_labor": worker.is_blocked_for_labor,
        },
    )
    await _log_activity(
        db,
        current_user.id,
        "worker_labor_control",
        f"근로자 노무 통제: worker_id={worker.id}, action={action}",
    )

    return APIResponse.ok(
        {
            "id": str(worker.id),
            "is_blocked_for_labor": worker.is_blocked_for_labor,
            "block_reason": worker.block_reason,
            "blocked_by_user_id": str(worker.blocked_by_user_id) if worker.blocked_by_user_id else None,
            "blocked_at": _to_iso(worker.blocked_at),
        }
    )


@router.get("/labor/worker-documents/{worker_document_id}/download")
async def download_worker_document(
    worker_document_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    _require_worker_document_upload_role(current_user)

    document = (
        await db.execute(select(WorkerDocument).where(WorkerDocument.id == worker_document_id))
    ).scalar_one_or_none()
    if not document:
        raise NotFoundException("worker_document", worker_document_id)

    worker = (
        await db.execute(select(DailyWorker).where(DailyWorker.id == document.worker_id))
    ).scalar_one_or_none()
    if not worker:
        raise NotFoundException("daily_worker", document.worker_id)
    if current_user.organization_id is not None and worker.organization_id != current_user.organization_id:
        raise NotFoundException("worker_document", worker_document_id)

    if not document.storage_path:
        raise HTTPException(status_code=404, detail="업로드된 파일이 없어요.")

    absolute_path = storage_service.get_absolute_path(document.storage_path)
    if not absolute_path.exists():
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없어요.")

    content = absolute_path.read_bytes()
    filename = document.original_filename or absolute_path.name
    encoded_filename = quote(filename, safe="")

    return StreamingResponse(
        io.BytesIO(content),
        media_type=document.mime_type or "application/octet-stream",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
        },
    )


class WorkRecordInput(BaseModel):
    worker_id: str
    project_id: str
    work_date: str
    man_days: Decimal


class WorkRecordBatchRequest(BaseModel):
    records: list[WorkRecordInput]


@router.get("/labor/work-records", response_model=APIResponse[list[dict]])
async def get_work_records(
    db: DBSession,
    current_user: CurrentUser,
    project_id: str,
    year: int,
    month: int,
):
    project_int = _to_int(project_id)
    await _get_project_for_user(db, project_int, current_user)

    month_start = date(year, month, 1)
    month_end = date(year + (month // 12), (month % 12) + 1, 1)

    result = await db.execute(
        select(WorkRecord)
        .where(WorkRecord.project_id == project_int)
        .where(WorkRecord.work_date >= month_start)
        .where(WorkRecord.work_date < month_end)
    )
    rows = result.scalars().all()

    return APIResponse.ok([
        {
            "id": str(r.id),
            "worker_id": str(r.worker_id),
            "project_id": str(r.project_id),
            "work_date": r.work_date.isoformat(),
            "man_days": float(r.man_days),
        }
        for r in rows
    ])


@router.post("/labor/work-records/batch", response_model=APIResponse[dict])
async def upsert_work_records(
    payload: WorkRecordBatchRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    _require_non_super_admin_for_labor_actions(current_user)

    if settings.labor_deployment_gate_enabled and payload.records:
        worker_ids = sorted({_to_int(record.worker_id) for record in payload.records})
        worker_query = select(DailyWorker).where(DailyWorker.id.in_(worker_ids))
        if current_user.organization_id is not None:
            worker_query = worker_query.where(DailyWorker.organization_id == current_user.organization_id)

        workers = (await db.execute(worker_query)).scalars().all()
        workers_by_id = {worker.id: worker for worker in workers}

        unknown_worker_ids = [worker_id for worker_id in worker_ids if worker_id not in workers_by_id]
        if unknown_worker_ids:
            raise HTTPException(
                status_code=404,
                detail={
                    "message": "일부 근로자를 찾을 수 없어요.",
                    "missing_worker_ids": unknown_worker_ids,
                },
            )

        blocked_workers: list[dict] = []
        for worker in workers:
            missing_requirements: list[str] = []
            if not worker.has_id_card:
                missing_requirements.append("id_card")
            if not worker.has_safety_cert:
                missing_requirements.append("safety_cert")
            if worker.is_blocked_for_labor:
                missing_requirements.append("labor_blocked")
            has_consent = await _has_consent_record_for_worker(db, worker)
            if not has_consent:
                missing_requirements.append("consent_record")

            if missing_requirements:
                blocked_workers.append(
                    {
                        "worker_id": str(worker.id),
                        "worker_name": worker.name,
                        "missing_requirements": missing_requirements,
                        "block_reason": worker.block_reason,
                    }
                )

        if blocked_workers:
            raise HTTPException(
                status_code=422,
                detail={
                    "message": "필수 서류/동의가 완료되지 않은 근로자가 있어 투입 기록을 저장할 수 없어요.",
                    "blocked_workers": blocked_workers,
                },
            )

    updated = 0

    for record in payload.records:
        worker_id = _to_int(record.worker_id)
        project_id = _to_int(record.project_id)

        await _get_project_for_user(db, project_id, current_user)

        work_date = date.fromisoformat(record.work_date)
        existing = await db.execute(
            select(WorkRecord)
            .where(WorkRecord.worker_id == worker_id)
            .where(WorkRecord.project_id == project_id)
            .where(WorkRecord.work_date == work_date)
        )
        row = existing.scalar_one_or_none()

        if row:
            row.man_days = record.man_days
            row.updated_at = datetime.utcnow()
        else:
            db.add(
                WorkRecord(
                    worker_id=worker_id,
                    project_id=project_id,
                    work_date=work_date,
                    man_days=record.man_days,
                )
            )
        updated += 1

    return APIResponse.ok({"updated_count": updated})


@router.delete("/labor/work-records/{work_record_id}", response_model=APIResponse[dict])
async def delete_work_record(
    work_record_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    _require_non_super_admin_for_labor_actions(current_user)
    result = await db.execute(select(WorkRecord).where(WorkRecord.id == work_record_id))
    row = result.scalar_one_or_none()
    if not row:
        raise NotFoundException("work_record", work_record_id)

    await _get_project_for_user(db, row.project_id, current_user)
    await db.delete(row)

    return APIResponse.ok({"id": str(work_record_id)})


_REPORT_TOTAL_KEYS = (
    "total_labor_cost",
    "total_income_tax",
    "total_resident_tax",
    "total_health_insurance",
    "total_longterm_care",
    "total_national_pension",
    "total_employment_insurance",
    "total_deductions",
    "total_net_pay",
)


def _empty_report_totals() -> dict[str, int]:
    return {key: 0 for key in _REPORT_TOTAL_KEYS}


def _month_bounds(year: int, month: int) -> tuple[date, date]:
    if month < 1 or month > 12:
        raise HTTPException(status_code=422, detail="month는 1~12 범위여야 해요")
    month_start = date(year, month, 1)
    month_end = date(year + (month // 12), (month % 12) + 1, 1)
    return month_start, month_end


def _parse_month_param(month: str) -> tuple[int, int]:
    try:
        year_str, month_str = month.split("-", 1)
        year = int(year_str)
        month_int = int(month_str)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail="month는 YYYY-MM 형식이어야 해요") from exc
    _month_bounds(year, month_int)
    return year, month_int


def _mask_worker_identifier(worker: DailyWorker) -> str:
    birth_digits = "".join(ch for ch in worker.birth_date if ch.isdigit())[:6]
    if len(birth_digits) < 6:
        birth_digits = "000000"
    gender_digit = str(worker.gender or 1)[0]
    return f"{birth_digits}-{gender_digit}******"


def _build_work_days(records: list[WorkRecord], cap_per_day: Optional[float] = None) -> dict[int, float]:
    out: dict[int, float] = {}
    for row in records:
        day = row.work_date.day
        current = out.get(day, 0.0)
        updated = current + float(row.man_days)
        if cap_per_day is not None:
            updated = min(updated, cap_per_day)
        out[day] = round(updated, 2)
    return dict(sorted(out.items(), key=lambda item: item[0]))


def _build_worker_entry(
    worker: DailyWorker,
    work_days: dict[int, float],
    rates: InsuranceRate,
    year: int,
    month: int,
) -> Optional[dict]:
    total_man_days = sum(v for v in work_days.values() if v > 0)
    if total_man_days <= 0:
        return None

    total_days = len([v for v in work_days.values() if v > 0])
    labor_cost = int((Decimal(worker.daily_rate) * Decimal(str(total_man_days))).quantize(Decimal("1"), rounding=ROUND_DOWN))
    deductions = _calc_worker_deductions(labor_cost, rates)

    last_work_day = 0
    for day in range(31, 0, -1):
        if (work_days.get(day) or 0) > 0:
            last_work_day = day
            break

    month_str = f"{year}{month:02d}"
    nts_last_work_date = f"{month_str}{last_work_day:02d}" if last_work_day > 0 else ""

    return {
        "worker_id": str(worker.id),
        "worker_name": worker.name,
        "job_type": worker.job_type,
        "job_type_code": worker.job_type_code or "",
        "team": worker.team,
        "ssn_masked": _mask_worker_identifier(worker),
        "ssn_full": "",
        "daily_rate": int(worker.daily_rate),
        "is_foreign": worker.is_foreign,
        "nationality_code": worker.nationality_code or "",
        "visa_status": worker.visa_status or "",
        "english_name": worker.english_name or "",
        "phone": worker.phone or "",
        "insurance_type": "5",
        "work_days": work_days,
        "total_days": total_days,
        "total_man_days": float(total_man_days),
        "total_labor_cost": labor_cost,
        "nontaxable_income": 0,
        "nts_pay_month": month_str,
        "nts_work_month": month_str,
        "nts_last_work_date": nts_last_work_date,
        **deductions,
    }


def _apply_entry_to_totals(entry: dict, totals: dict[str, int]) -> None:
    totals["total_labor_cost"] += int(entry["total_labor_cost"])
    totals["total_income_tax"] += int(entry["income_tax"])
    totals["total_resident_tax"] += int(entry["resident_tax"])
    totals["total_health_insurance"] += int(entry["health_insurance"])
    totals["total_longterm_care"] += int(entry["longterm_care"])
    totals["total_national_pension"] += int(entry["national_pension"])
    totals["total_employment_insurance"] += int(entry["employment_insurance"])
    totals["total_deductions"] += int(entry["total_deductions"])
    totals["total_net_pay"] += int(entry["net_pay"])


async def _build_site_report(
    db: AsyncSession,
    organization_id: int,
    project_id: int,
    year: int,
    month: int,
) -> dict:
    project = (
        await db.execute(
            select(Project)
            .where(Project.id == project_id)
            .where(Project.organization_id == organization_id)
        )
    ).scalar_one_or_none()
    if not project:
        raise NotFoundException("project", project_id)

    month_start, month_end = _month_bounds(year, month)

    work_rows = (
        await db.execute(
            select(WorkRecord)
            .where(WorkRecord.project_id == project_id)
            .where(WorkRecord.work_date >= month_start)
            .where(WorkRecord.work_date < month_end)
        )
    ).scalars().all()

    by_worker: dict[int, list[WorkRecord]] = {}
    for row in work_rows:
        by_worker.setdefault(row.worker_id, []).append(row)

    worker_ids = list(by_worker.keys())
    workers_by_id: dict[int, DailyWorker] = {}
    if worker_ids:
        worker_rows = (
            await db.execute(
                select(DailyWorker)
                .where(DailyWorker.organization_id == organization_id)
                .where(DailyWorker.id.in_(worker_ids))
            )
        ).scalars().all()
        workers_by_id = {worker.id: worker for worker in worker_rows}

    rates = await _get_or_create_insurance_rate(db, organization_id, year)
    totals = _empty_report_totals()
    entries: list[dict] = []

    sorted_worker_ids = sorted(
        by_worker.keys(),
        key=lambda worker_id: workers_by_id.get(worker_id).name if workers_by_id.get(worker_id) else "",
    )
    for worker_id in sorted_worker_ids:
        worker = workers_by_id.get(worker_id)
        if not worker:
            continue
        work_days = _build_work_days(by_worker[worker_id])
        entry = _build_worker_entry(worker, work_days, rates, year, month)
        if not entry:
            continue
        entries.append(entry)
        _apply_entry_to_totals(entry, totals)

    org = (await db.execute(select(Organization).where(Organization.id == organization_id))).scalar_one_or_none()
    return {
        "project_id": str(project.id),
        "project_name": project.name,
        "year": year,
        "month": month,
        "organization_name": org.name if org else "",
        "entries": entries,
        "totals": totals,
    }


async def _build_monthly_consolidated_report(
    db: AsyncSession,
    organization_id: int,
    year: int,
    month: int,
) -> dict:
    month_start, month_end = _month_bounds(year, month)

    project_rows = (
        await db.execute(
            select(Project)
            .where(Project.organization_id == organization_id)
            .order_by(Project.name.asc())
        )
    ).scalars().all()

    work_rows = (
        await db.execute(
            select(WorkRecord)
            .join(Project, WorkRecord.project_id == Project.id)
            .where(Project.organization_id == organization_id)
            .where(WorkRecord.work_date >= month_start)
            .where(WorkRecord.work_date < month_end)
        )
    ).scalars().all()

    by_worker: dict[int, list[WorkRecord]] = {}
    included_project_ids: set[int] = set()
    for row in work_rows:
        by_worker.setdefault(row.worker_id, []).append(row)
        included_project_ids.add(row.project_id)

    worker_ids = list(by_worker.keys())
    workers_by_id: dict[int, DailyWorker] = {}
    if worker_ids:
        worker_rows = (
            await db.execute(
                select(DailyWorker)
                .where(DailyWorker.organization_id == organization_id)
                .where(DailyWorker.id.in_(worker_ids))
            )
        ).scalars().all()
        workers_by_id = {worker.id: worker for worker in worker_rows}

    rates = await _get_or_create_insurance_rate(db, organization_id, year)
    totals = _empty_report_totals()
    entries: list[dict] = []

    sorted_worker_ids = sorted(
        by_worker.keys(),
        key=lambda worker_id: workers_by_id.get(worker_id).name if workers_by_id.get(worker_id) else "",
    )
    for worker_id in sorted_worker_ids:
        worker = workers_by_id.get(worker_id)
        if not worker:
            continue
        # Consolidated monthly report merges duplicate working days across sites.
        work_days = _build_work_days(by_worker[worker_id], cap_per_day=1.0)
        entry = _build_worker_entry(worker, work_days, rates, year, month)
        if not entry:
            continue
        entries.append(entry)
        _apply_entry_to_totals(entry, totals)

    org = (await db.execute(select(Organization).where(Organization.id == organization_id))).scalar_one_or_none()
    included_projects = [p for p in project_rows if p.id in included_project_ids]
    return {
        "year": year,
        "month": month,
        "organization_name": org.name if org else "",
        "projects": [{"id": str(p.id), "name": p.name} for p in included_projects],
        "entries": entries,
        "totals": totals,
    }


def _serialize_insurance_rate(rate: InsuranceRate) -> dict:
    return {
        "id": str(rate.id),
        "effective_year": rate.effective_year,
        "effective_from": f"{rate.effective_year}-01-01",
        "income_deduction": float(rate.income_deduction),
        "simplified_tax_rate": float(rate.simplified_tax_rate),
        "local_tax_rate": float(rate.local_tax_rate),
        "employment_insurance_rate": float(rate.employment_insurance_rate),
        "health_insurance_rate": float(rate.health_insurance_rate),
        "longterm_care_rate": float(rate.longterm_care_rate),
        "national_pension_rate": float(rate.national_pension_rate),
        "pension_upper_limit": float(rate.pension_upper_limit),
        "pension_lower_limit": float(rate.pension_lower_limit),
        "health_premium_upper": float(rate.health_premium_upper),
        "health_premium_lower": float(rate.health_premium_lower),
    }


def _sanitize_filename(value: str) -> str:
    stripped = value.strip()
    if not stripped:
        return "report"
    return "".join("_" if ch in '\\/:*?"<>|' else ch for ch in stripped)


def _xlsx_response(xlsx_bytes: bytes, filename: str) -> StreamingResponse:
    encoded_filename = quote(filename)
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
        },
    )


def _apply_header_style(cell) -> None:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_payroll_workbook(report: dict, title: str, include_project_list: bool = False) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "명세서"

    headers = [
        "No.",
        "성명",
        "직종",
        "주민번호",
        "단가",
        *[str(day) for day in range(1, 32)],
        "출력일수",
        "공수",
        "노무비",
        "갑근세",
        "주민세",
        "건강보험",
        "요양보험",
        "국민연금",
        "고용보험",
        "공제계",
        "차감지급액",
    ]

    worksheet.cell(row=1, column=1, value=title)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    worksheet.cell(row=1, column=1).font = Font(size=14, bold=True)
    worksheet.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    info_row = 3
    worksheet.cell(info_row, 1, "회사명")
    worksheet.cell(info_row, 2, report.get("organization_name", ""))
    if report.get("project_name"):
        info_row += 1
        worksheet.cell(info_row, 1, "현장명")
        worksheet.cell(info_row, 2, report.get("project_name", ""))
    info_row += 1
    worksheet.cell(info_row, 1, "기준월")
    worksheet.cell(info_row, 2, f"{report['year']}년 {report['month']}월")
    if include_project_list:
        project_names = ", ".join(project["name"] for project in report.get("projects", []))
        info_row += 1
        worksheet.cell(info_row, 1, "포함 현장")
        worksheet.cell(info_row, 2, project_names)

    header_row_index = info_row + 2
    worksheet.append([])
    worksheet.append(headers)
    for column_index in range(1, len(headers) + 1):
        _apply_header_style(worksheet.cell(row=header_row_index, column=column_index))

    start_data_row = header_row_index + 1
    entries = report.get("entries", [])
    for index, entry in enumerate(entries, start=1):
        row_data: list[object] = [
            index,
            entry["worker_name"],
            entry["job_type"],
            entry["ssn_masked"],
            entry["daily_rate"],
        ]
        for day in range(1, 32):
            man_days = entry["work_days"].get(day, 0)
            row_data.append(man_days if man_days > 0 else "")
        row_data.extend(
            [
                entry["total_days"],
                entry["total_man_days"],
                entry["total_labor_cost"],
                entry["income_tax"],
                entry["resident_tax"],
                entry["health_insurance"],
                entry["longterm_care"],
                entry["national_pension"],
                entry["employment_insurance"],
                entry["total_deductions"],
                entry["net_pay"],
            ]
        )
        worksheet.append(row_data)

    totals = report.get("totals", _empty_report_totals())
    totals_row = [
        "",
        "합계",
        "",
        "",
        "",
        *([""] * 31),
        "",
        "",
        totals["total_labor_cost"],
        totals["total_income_tax"],
        totals["total_resident_tax"],
        totals["total_health_insurance"],
        totals["total_longterm_care"],
        totals["total_national_pension"],
        totals["total_employment_insurance"],
        totals["total_deductions"],
        totals["total_net_pay"],
    ]
    worksheet.append(totals_row)
    totals_row_index = start_data_row + len(entries)
    for column_index in range(1, len(headers) + 1):
        _apply_header_style(worksheet.cell(row=totals_row_index, column=column_index))

    number_columns = [5, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47]
    for row in range(start_data_row, totals_row_index + 1):
        for col in number_columns:
            worksheet.cell(row=row, column=col).number_format = "#,##0.##"

    widths = {
        1: 6,
        2: 12,
        3: 12,
        4: 15,
        5: 10,
    }
    for col in range(6, 37):
        widths[col] = 4
    for col in range(37, 48):
        widths[col] = 11
    for col, width in widths.items():
        worksheet.column_dimensions[get_column_letter(col)].width = width

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _split_phone_number(phone: str) -> tuple[str, str, str]:
    digits = "".join(ch for ch in (phone or "") if ch.isdigit())
    if not digits:
        return "", "", ""
    if digits.startswith("02"):
        return "02", digits[2:6], digits[6:10]
    if len(digits) >= 11:
        return digits[:3], digits[3:7], digits[7:11]
    if len(digits) >= 10:
        return digits[:3], digits[3:6], digits[6:10]
    return digits[:3], digits[3:6], digits[6:]


def _validate_welfare_required_codes(entries: list[dict]) -> list[dict]:
    missing: list[dict] = []
    for entry in entries:
        missing_fields: list[str] = []
        invalid_values: dict[str, str] = {}

        job_type_code = normalize_code(entry.get("job_type_code"))
        if not job_type_code:
            missing_fields.append("job_type_code")
        elif not is_valid_job_type_code(job_type_code):
            invalid_values["job_type_code"] = job_type_code

        if entry.get("is_foreign"):
            nationality_code = normalize_code(entry.get("nationality_code"))
            visa_status = normalize_code(entry.get("visa_status"))

            if not nationality_code:
                missing_fields.append("nationality_code")
            elif not is_valid_nationality_code(nationality_code):
                invalid_values["nationality_code"] = nationality_code

            if not visa_status:
                missing_fields.append("visa_status")
            elif not is_valid_visa_status(visa_status):
                invalid_values["visa_status"] = visa_status

        if missing_fields or invalid_values:
            missing.append(
                {
                    "worker_id": entry["worker_id"],
                    "worker_name": entry["worker_name"],
                    "missing_fields": missing_fields,
                    "invalid_values": invalid_values,
                }
            )
    return missing


def _build_welfare_form_workbook(report: dict, template_version: str = "v1") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "★근로내역"

    headers = [
        "보험구분",
        "성명",
        "주민(외국인)등록번호",
        "국적코드",
        "체류자격코드",
        "전화(지역번호)",
        "전화(국번)",
        "전화(뒷번호)",
        "직종코드",
        *[f"{day}일" for day in range(1, 32)],
        "근로일수",
        "일평균근로시간",
        "보수지급기초일수",
        "보수총액(과세소득)",
        "임금총액",
        "이직사유코드",
        "보험료부과구분부호",
        "보험료부과구분사유",
        "국세청일용근로소득신고여부",
        "지급월",
        "총지급액(과세소득)",
        "비과세소득",
        "소득세",
        "지방소득세",
    ]
    sheet.append(headers)
    for column_index in range(1, len(headers) + 1):
        _apply_header_style(sheet.cell(row=1, column=column_index))

    pay_month = f"{report['year']}{report['month']:02d}"
    for entry in report.get("entries", []):
        area, mid, last = _split_phone_number(entry.get("phone", ""))
        nationality_code = entry.get("nationality_code") if entry.get("is_foreign") else "100"
        visa_status = entry.get("visa_status") if entry.get("is_foreign") else ""

        row_data: list[object] = [
            entry.get("insurance_type", "5"),
            entry["worker_name"],
            "",  # 주민등록번호는 수기 입력
            nationality_code or "",
            visa_status or "",
            area,
            mid,
            last,
            entry.get("job_type_code", ""),
        ]

        for day in range(1, 32):
            man_days = entry["work_days"].get(day, 0)
            row_data.append(man_days if man_days > 0 else 0)

        row_data.extend(
            [
                entry["total_days"],
                8,
                entry["total_days"],
                entry["total_labor_cost"],
                entry["total_labor_cost"],
                "",
                "",
                "",
                "Y" if entry["income_tax"] > 0 else "",
                entry.get("nts_pay_month", pay_month),
                entry["total_labor_cost"],
                entry.get("nontaxable_income", 0),
                entry["income_tax"],
                entry["resident_tax"],
            ]
        )
        sheet.append(row_data)

    for col in range(44, 55):
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=col).number_format = "#,##0"

    guide = workbook.create_sheet("작성안내")
    guide.append(["항목", "안내"])
    guide.append(["template_version", template_version])
    guide.append(["주민등록번호", "수기 입력 항목입니다."])
    guide.append(["주의", "직종/국적/체류자격 코드 누락 시 다운로드가 차단됩니다."])

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


@router.get("/labor/reports/site", response_model=APIResponse[dict])
async def generate_site_report(
    db: DBSession,
    current_user: CurrentUser,
    project_id: str,
    year: int,
    month: int,
):
    org_id = _require_org_id(current_user)
    project_int = _to_int(project_id)
    await _get_project_for_user(db, project_int, current_user)

    report = await _build_site_report(db, org_id, project_int, year, month)
    return APIResponse.ok(report)


@router.get("/labor/reports/consolidated", response_model=APIResponse[dict])
async def generate_consolidated_report(
    db: DBSession,
    current_user: CurrentUser,
    year: int,
    month: int,
):
    org_id = _require_org_id(current_user)
    report = await _build_monthly_consolidated_report(db, org_id, year, month)
    return APIResponse.ok(report)


@router.get("/reports/site-daily-labor")
async def download_site_daily_labor_report(
    db: DBSession,
    current_user: CurrentUser,
    project_id: str,
    month: str,
):
    org_id = _require_org_id(current_user)
    year, month_int = _parse_month_param(month)
    project_int = _to_int(project_id)
    await _get_project_for_user(db, project_int, current_user)

    report = await _build_site_report(db, org_id, project_int, year, month_int)
    xlsx_bytes = _build_payroll_workbook(
        report=report,
        title="현장별 일용신고명세서",
    )
    filename = f"{_sanitize_filename(report['project_name'])} 현장별 일용신고명세서({month_int}월).xlsx"
    return _xlsx_response(xlsx_bytes, filename)


@router.get("/reports/monthly-daily-labor")
async def download_monthly_daily_labor_report(
    db: DBSession,
    current_user: CurrentUser,
    month: str,
):
    org_id = _require_org_id(current_user)
    year, month_int = _parse_month_param(month)
    report = await _build_monthly_consolidated_report(db, org_id, year, month_int)
    xlsx_bytes = _build_payroll_workbook(
        report=report,
        title="월별 일용노무비 지급명세서",
        include_project_list=True,
    )
    filename = f"{_sanitize_filename(report['organization_name'])} {year}년 {month_int}월 일용근로자 사용 명단.xlsx"
    return _xlsx_response(xlsx_bytes, filename)


@router.get("/reports/welfare-form")
async def download_welfare_form(
    db: DBSession,
    current_user: CurrentUser,
    project_id: str,
    month: str,
    template_version: str = Query(default="v1"),
):
    org_id = _require_org_id(current_user)
    year, month_int = _parse_month_param(month)
    project_int = _to_int(project_id)
    await _get_project_for_user(db, project_int, current_user)

    report = await _build_site_report(db, org_id, project_int, year, month_int)
    missing_codes = _validate_welfare_required_codes(report["entries"])
    if missing_codes:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "근로복지공단 코드 매핑 누락/오류로 파일을 생성할 수 없어요.",
                "missing": missing_codes,
            },
        )

    xlsx_bytes = _build_welfare_form_workbook(report, template_version=template_version)
    filename = (
        f"{_sanitize_filename(report['project_name'])} "
        f"근로복지공단 신고 양식({year}년 {month_int}월).xlsx"
    )
    return _xlsx_response(xlsx_bytes, filename)


class InsuranceRateUpsertRequest(BaseModel):
    effective_year: int
    income_deduction: Decimal
    simplified_tax_rate: Decimal
    local_tax_rate: Decimal
    employment_insurance_rate: Decimal
    health_insurance_rate: Decimal
    longterm_care_rate: Decimal
    national_pension_rate: Decimal
    pension_upper_limit: Decimal
    pension_lower_limit: Decimal
    health_premium_upper: Decimal
    health_premium_lower: Decimal


class AdminRateUpdateRequest(BaseModel):
    income_deduction: Decimal
    simplified_tax_rate: Decimal
    local_tax_rate: Decimal
    employment_insurance_rate: Decimal
    health_insurance_rate: Decimal
    longterm_care_rate: Decimal
    national_pension_rate: Decimal
    pension_upper_limit: Decimal
    pension_lower_limit: Decimal
    health_premium_upper: Decimal
    health_premium_lower: Decimal
    effective_year: Optional[int] = None
    effective_from: Optional[date] = None


@router.get("/labor/insurance-rates", response_model=APIResponse[list[dict]])
async def get_insurance_rates(
    db: DBSession,
    current_user: CurrentUser,
    year: Optional[int] = None,
):
    org_id = _require_org_id(current_user)

    query = select(InsuranceRate).where(InsuranceRate.organization_id == org_id)
    if year:
        query = query.where(InsuranceRate.effective_year == year)

    query = query.order_by(InsuranceRate.effective_year.desc())
    rows = (await db.execute(query)).scalars().all()
    return APIResponse.ok([_serialize_insurance_rate(row) for row in rows])


@router.post("/labor/insurance-rates", response_model=APIResponse[dict])
async def create_insurance_rate(
    payload: InsuranceRateUpsertRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    _require_non_super_admin_for_labor_actions(current_user)
    org_id = _require_org_id(current_user)

    existing = (
        await db.execute(
            select(InsuranceRate)
            .where(InsuranceRate.organization_id == org_id)
            .where(InsuranceRate.effective_year == payload.effective_year)
        )
    ).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=409, detail="해당 연도의 요율이 이미 등록되어 있어요")

    rate = InsuranceRate(
        organization_id=org_id,
        effective_year=payload.effective_year,
        income_deduction=payload.income_deduction,
        simplified_tax_rate=payload.simplified_tax_rate,
        local_tax_rate=payload.local_tax_rate,
        employment_insurance_rate=payload.employment_insurance_rate,
        health_insurance_rate=payload.health_insurance_rate,
        longterm_care_rate=payload.longterm_care_rate,
        national_pension_rate=payload.national_pension_rate,
        pension_upper_limit=payload.pension_upper_limit,
        pension_lower_limit=payload.pension_lower_limit,
        health_premium_upper=payload.health_premium_upper,
        health_premium_lower=payload.health_premium_lower,
    )
    db.add(rate)
    await db.flush()

    return APIResponse.ok({"id": str(rate.id)})


@router.patch("/labor/insurance-rates/{rate_id}", response_model=APIResponse[dict])
async def update_insurance_rate(
    rate_id: int,
    payload: InsuranceRateUpsertRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    _require_non_super_admin_for_labor_actions(current_user)
    org_id = _require_org_id(current_user)

    rate = (
        await db.execute(
            select(InsuranceRate)
            .where(InsuranceRate.id == rate_id)
            .where(InsuranceRate.organization_id == org_id)
        )
    ).scalar_one_or_none()
    if not rate:
        raise NotFoundException("insurance_rate", rate_id)

    duplicate_year = (
        await db.execute(
            select(InsuranceRate)
            .where(InsuranceRate.organization_id == org_id)
            .where(InsuranceRate.effective_year == payload.effective_year)
            .where(InsuranceRate.id != rate.id)
        )
    ).scalar_one_or_none()
    if duplicate_year:
        raise HTTPException(status_code=409, detail="같은 적용연도 요율이 이미 존재해요")

    rate.effective_year = payload.effective_year
    rate.income_deduction = payload.income_deduction
    rate.simplified_tax_rate = payload.simplified_tax_rate
    rate.local_tax_rate = payload.local_tax_rate
    rate.employment_insurance_rate = payload.employment_insurance_rate
    rate.health_insurance_rate = payload.health_insurance_rate
    rate.longterm_care_rate = payload.longterm_care_rate
    rate.national_pension_rate = payload.national_pension_rate
    rate.pension_upper_limit = payload.pension_upper_limit
    rate.pension_lower_limit = payload.pension_lower_limit
    rate.health_premium_upper = payload.health_premium_upper
    rate.health_premium_lower = payload.health_premium_lower
    rate.updated_at = datetime.utcnow()

    return APIResponse.ok({"id": str(rate.id)})


@router.get("/admin/rates", response_model=APIResponse[dict])
async def get_admin_rates(
    db: DBSession,
    current_user: CurrentUser,
    as_of: Optional[date] = Query(default=None),
):
    org_id = _require_org_id(current_user)
    as_of_date = as_of or date.today()

    rows = (
        await db.execute(
            select(InsuranceRate)
            .where(InsuranceRate.organization_id == org_id)
            .order_by(InsuranceRate.effective_year.desc())
        )
    ).scalars().all()

    active_rate = next((row for row in rows if row.effective_year <= as_of_date.year), None)
    if active_rate is None and rows:
        active_rate = rows[0]

    return APIResponse.ok(
        {
            "as_of": as_of_date.isoformat(),
            "active_rate": _serialize_insurance_rate(active_rate) if active_rate else None,
            "versions": [_serialize_insurance_rate(row) for row in rows],
        }
    )


@router.put("/admin/rates/{rate_version_id}", response_model=APIResponse[dict])
async def update_admin_rate(
    rate_version_id: int,
    payload: AdminRateUpdateRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    org_id = _require_org_id(current_user)

    rate = (
        await db.execute(
            select(InsuranceRate)
            .where(InsuranceRate.id == rate_version_id)
            .where(InsuranceRate.organization_id == org_id)
        )
    ).scalar_one_or_none()
    if not rate:
        raise NotFoundException("insurance_rate", rate_version_id)

    if payload.effective_from and payload.effective_year and payload.effective_from.year != payload.effective_year:
        raise HTTPException(status_code=400, detail="effective_from과 effective_year가 일치해야 해요")

    next_effective_year = (
        payload.effective_year
        if payload.effective_year is not None
        else payload.effective_from.year if payload.effective_from else rate.effective_year
    )

    duplicate_year = (
        await db.execute(
            select(InsuranceRate)
            .where(InsuranceRate.organization_id == org_id)
            .where(InsuranceRate.effective_year == next_effective_year)
            .where(InsuranceRate.id != rate.id)
        )
    ).scalar_one_or_none()
    if duplicate_year:
        raise HTTPException(status_code=409, detail="같은 적용연도 요율이 이미 존재해요")

    rate.effective_year = next_effective_year
    rate.income_deduction = payload.income_deduction
    rate.simplified_tax_rate = payload.simplified_tax_rate
    rate.local_tax_rate = payload.local_tax_rate
    rate.employment_insurance_rate = payload.employment_insurance_rate
    rate.health_insurance_rate = payload.health_insurance_rate
    rate.longterm_care_rate = payload.longterm_care_rate
    rate.national_pension_rate = payload.national_pension_rate
    rate.pension_upper_limit = payload.pension_upper_limit
    rate.pension_lower_limit = payload.pension_lower_limit
    rate.health_premium_upper = payload.health_premium_upper
    rate.health_premium_lower = payload.health_premium_lower
    rate.updated_at = datetime.utcnow()

    await _log_activity(
        db=db,
        user_id=current_user.id,
        action="settings_change",
        description=f"보험요율 버전({rate.id}) 수정",
    )

    return APIResponse.ok(
        {
            "id": str(rate.id),
            "effective_year": rate.effective_year,
            "effective_from": f"{rate.effective_year}-01-01",
        }
    )


# ----------------------------
# Worker app and notifications
# ----------------------------


class WorkerAccessRequestBody(BaseModel):
    phone: str


class WorkerVerifyRequestBody(BaseModel):
    request_id: str
    code: str


class WorkerInviteVerifyRequestBody(BaseModel):
    invite_token: str


@router.post("/workers/access", response_model=APIResponse[dict])
async def request_worker_access(
    payload: WorkerAccessRequestBody,
    db: DBSession,
):
    code = f"{random.randint(0, 999999):06d}"
    record = WorkerAccessRequest(
        phone=payload.phone,
        code=code,
        expires_at=datetime.utcnow() + timedelta(minutes=10),
    )
    db.add(record)
    await db.flush()

    return APIResponse.ok({"request_id": str(record.id)})


@router.post("/workers/verify", response_model=APIResponse[dict])
async def verify_worker_access(
    payload: WorkerVerifyRequestBody,
    db: DBSession,
):
    request_id = _to_int(payload.request_id)
    record = (await db.execute(select(WorkerAccessRequest).where(WorkerAccessRequest.id == request_id))).scalar_one_or_none()
    if not record:
        raise NotFoundException("worker_access_request", payload.request_id)

    if record.expires_at < datetime.utcnow() or record.code != payload.code:
        raise HTTPException(status_code=400, detail="인증번호를 다시 확인해 주세요")

    record.verified = True

    user_result = await db.execute(select(User).where(User.phone == record.phone, User.role == UserRole.WORKER))
    worker_user = user_result.scalar_one_or_none()

    if not worker_user:
        username = f"worker_{record.phone.replace('-', '')}"
        worker_user = User(
            username=username,
            name="현장근로자",
            phone=record.phone,
            email=None,
            role=UserRole.WORKER,
            organization_id=None,
            password_hash=get_password_hash(record.phone.replace("-", "")[-4:] or "1234"),
            is_active=True,
        )
        db.add(worker_user)
        await db.flush()

    record.worker_id = worker_user.id

    return APIResponse.ok({"worker_id": str(worker_user.id)})


@router.post("/workers/invite/verify", response_model=APIResponse[dict])
async def verify_worker_invite(
    payload: WorkerInviteVerifyRequestBody,
    db: DBSession,
):
    username = f"worker_invite_{payload.invite_token[-8:]}"
    result = await db.execute(select(User).where(User.username == username))
    user = result.scalar_one_or_none()
    if not user:
        user = User(
            username=username,
            name="초대 근로자",
            phone=None,
            email=None,
            role=UserRole.WORKER,
            organization_id=None,
            password_hash=get_password_hash("123456"),
            is_active=True,
        )
        db.add(user)
        await db.flush()

    return APIResponse.ok({"worker_id": str(user.id)})


class WorkerSignRequest(BaseModel):
    signature_data: str


@router.get("/workers/contracts/{contract_id}", response_model=APIResponse[dict])
async def get_worker_contract(
    contract_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    labor_contract = await _ensure_labor_contract_stub(
        db,
        contract_id,
        worker_phone=current_user.phone,
    )
    _require_worker_contract_access(labor_contract, current_user)

    project = (await db.execute(select(Project).where(Project.id == labor_contract.project_id))).scalar_one_or_none()

    content = (
        f"본 계약은 {project.name if project else '현장'}에서 {labor_contract.work_date.isoformat()}에 수행되는 "
        f"{labor_contract.work_type or '일용 근로'} 업무에 대한 계약입니다."
    )

    status_value = "signed" if labor_contract.status in {LaborContractStatus.SIGNED, LaborContractStatus.PAID} else "pending"

    return APIResponse.ok(
        {
            "id": str(labor_contract.id),
            "project_name": project.name if project else "",
            "work_date": labor_contract.work_date.isoformat(),
            "role": labor_contract.work_type or "근로자",
            "daily_rate": int(labor_contract.daily_rate),
            "status": status_value,
            "content": content,
        }
    )


@router.post("/workers/contracts/{contract_id}/sign", response_model=APIResponse[dict])
async def sign_worker_contract(
    contract_id: int,
    payload: WorkerSignRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    labor_contract = await _ensure_labor_contract_stub(
        db,
        contract_id,
        worker_phone=current_user.phone,
    )
    _require_worker_contract_access(labor_contract, current_user)

    labor_contract.status = LaborContractStatus.SIGNED
    labor_contract.signed_at = datetime.utcnow()
    labor_contract.worker_signature_path = f"signatures/workers/{contract_id}.png"

    return APIResponse.ok({"id": str(labor_contract.id), "status": "signed", "signed_at": labor_contract.signed_at.isoformat()})


@router.get("/workers/{worker_id}/paystubs", response_model=APIResponse[list[dict]])
async def get_worker_paystubs(
    worker_id: str,
    db: DBSession,
    current_user: CurrentUser,
):
    worker_user_id = await _require_worker_self(db, current_user, worker_id)
    await _seed_paystub_if_empty(db, worker_user_id)

    rows = (await db.execute(
        select(Paystub)
        .where(Paystub.worker_id == worker_user_id)
        .order_by(Paystub.created_at.desc())
    )).scalars().all()

    return APIResponse.ok([
        {
            "id": str(r.id),
            "month": r.month,
            "amount": r.net_amount,
            "status": r.status.value,
            "date": r.date,
        }
        for r in rows
    ])


@router.get("/workers/{worker_id}/paystubs/{paystub_id}", response_model=APIResponse[dict])
async def get_worker_paystub(
    worker_id: str,
    paystub_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    worker_user_id = await _require_worker_self(db, current_user, worker_id)

    paystub = (await db.execute(
        select(Paystub)
        .where(Paystub.id == paystub_id)
        .where(Paystub.worker_id == worker_user_id)
    )).scalar_one_or_none()
    if not paystub:
        raise NotFoundException("paystub", paystub_id)

    items = (await db.execute(
        select(PaystubItem).where(PaystubItem.paystub_id == paystub.id)
    )).scalars().all()

    return APIResponse.ok(
        {
            "id": str(paystub.id),
            "title": paystub.title,
            "total_amount": paystub.total_amount,
            "deductions": paystub.deductions,
            "net_amount": paystub.net_amount,
            "items": [{"label": item.label, "amount": item.amount} for item in items],
            "status": paystub.status.value,
        }
    )


@router.post("/workers/{worker_id}/paystubs/{paystub_id}/ack", response_model=APIResponse[dict])
async def ack_worker_paystub(
    worker_id: str,
    paystub_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    worker_user_id = await _require_worker_self(db, current_user, worker_id)

    paystub = (await db.execute(
        select(Paystub)
        .where(Paystub.id == paystub_id)
        .where(Paystub.worker_id == worker_user_id)
    )).scalar_one_or_none()
    if not paystub:
        raise NotFoundException("paystub", paystub_id)

    paystub.status = PaystubStatus.CONFIRMED
    return APIResponse.ok({"received_at": datetime.utcnow().isoformat()})


@router.get("/workers/{worker_id}/profile", response_model=APIResponse[dict])
async def get_worker_profile(
    worker_id: str,
    db: DBSession,
    current_user: CurrentUser,
):
    worker_user_id = await _require_worker_self(db, current_user, worker_id)

    user = (await db.execute(select(User).where(User.id == worker_user_id))).scalar_one_or_none()
    if not user:
        raise NotFoundException("user", worker_id)

    docs = await _seed_worker_documents(db, worker_user_id)

    return APIResponse.ok(
        {
            "id": str(user.id),
            "name": user.name,
            "role": "근로자",
            "documents": [
                {
                    "id": d.document_id,
                    "name": d.name,
                    "status": "submitted" if d.status == "submitted" else "pending",
                }
                for d in docs
            ],
        }
    )


@router.post("/workers/{worker_id}/documents/{doc_id}", response_model=APIResponse[dict])
async def upload_worker_document(
    worker_id: str,
    doc_id: str,
    db: DBSession,
    current_user: CurrentUser,
    file: UploadFile = File(...),
):
    worker_user_id = await _require_worker_self(db, current_user, worker_id)
    normalized_doc_type = _normalize_worker_document_type(doc_id)
    ext, mime_type = _validate_worker_document_file(file)

    docs = await _seed_worker_documents(db, worker_user_id)
    target = next((d for d in docs if d.document_id == doc_id), None)
    if not target:
        target = WorkerDocument(
            worker_id=worker_user_id,
            document_id=doc_id,
            document_type=normalized_doc_type,
            name=doc_id,
            status="pending",
            review_status="pending_review",
        )
        db.add(target)
        await db.flush()

    content = await file.read()
    await file.seek(0)
    if len(content) == 0:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="빈 파일은 업로드할 수 없어요.",
        )
    file_hash = hashlib.sha256(content).hexdigest()
    storage_path = await storage_service.save_bytes(
        data=content,
        category="contracts",
        subfolder=f"workers/{worker_user_id}",
        filename=f"{doc_id}_{int(datetime.utcnow().timestamp())}{ext}",
    )

    target.document_type = normalized_doc_type
    target.storage_path = storage_path
    target.status = "submitted"
    target.original_filename = file.filename or f"{doc_id}{ext}"
    target.mime_type = mime_type or None
    target.file_size_bytes = len(content)
    target.file_hash_sha256 = file_hash
    target.review_status = "pending_review"
    target.review_reason = None
    target.reviewed_by_user_id = None
    target.reviewed_at = None
    target.anomaly_flags = []
    target.uploaded_at = datetime.utcnow()

    return APIResponse.ok(
        {
            "id": str(target.id),
            "document_id": target.document_id,
            "storage_path": target.storage_path,
            "status": "submitted",
            "uploaded_at": target.uploaded_at.isoformat() if target.uploaded_at else datetime.utcnow().isoformat(),
        }
    )


@router.get("/notifications", response_model=APIResponse[list[dict]])
async def get_notifications(
    db: DBSession,
):
    rows = (await db.execute(
        select(AppNotification)
        .order_by(AppNotification.created_at.desc())
        .limit(20)
    )).scalars().all()

    if not rows:
        now = datetime.utcnow()
        defaults = [
            AppNotification(type=NotificationType.NOTICE, title="시스템 안내", message="새로운 공지사항이 있습니다.", time=now.strftime("%H:%M"), read=False),
            AppNotification(type=NotificationType.CONTRACT, title="계약 알림", message="근로계약서 서명 요청이 도착했습니다.", time=(now - timedelta(hours=2)).strftime("%H:%M"), read=False),
        ]
        for item in defaults:
            db.add(item)
        await db.flush()
        rows = defaults

    return APIResponse.ok(
        [
            {
                "id": str(n.id),
                "type": n.type.value,
                "title": n.title,
                "message": n.message,
                "time": n.time,
                "read": n.read,
            }
            for n in rows
        ]
    )


@router.post("/notifications/{notification_id}/read", response_model=APIResponse[dict])
async def mark_notification_read(
    notification_id: int,
    db: DBSession,
):
    notification = (await db.execute(select(AppNotification).where(AppNotification.id == notification_id))).scalar_one_or_none()
    if not notification:
        raise NotFoundException("notification", notification_id)

    notification.read = True
    return APIResponse.ok({"id": str(notification.id), "read": notification.read})


# ----------------------------
# Billing / SA
# ----------------------------


class PaymentMethodRequest(BaseModel):
    card_number: str
    expiry: str


class CustomTrialRequest(BaseModel):
    end_date: str
    reason: Optional[str] = None


@router.get("/billing/overview", response_model=APIResponse[dict])
async def get_billing_overview(
    db: DBSession,
    current_user: CurrentUser,
):
    org_id = _require_org_id(current_user)

    subscription = (await db.execute(select(Subscription).where(Subscription.organization_id == org_id))).scalar_one_or_none()
    if not subscription:
        return APIResponse.ok(
            {
                "plan": "",
                "subscription_start_date": None,
                "subscription_end_date": "",
                "days_remaining": 0,
                "is_custom_trial": True,
                "billing_amount": 0,
                "seats_used": 0,
                "seats_total": 0,
                "history": [],
            }
        )

    payments = (await db.execute(
        select(Payment)
        .where(Payment.organization_id == org_id)
        .order_by(Payment.created_at.desc())
        .limit(20)
    )).scalars().all()

    plan_label = {
        SubscriptionPlan.STARTER: "STARTER",
        SubscriptionPlan.STANDARD: "STANDARD",
        SubscriptionPlan.PREMIUM: "PREMIUM",
    }.get(subscription.plan, subscription.plan.value)

    now = datetime.utcnow()
    days_remaining = max(0, (subscription.expires_at.date() - now.date()).days)

    return APIResponse.ok(
        {
            "plan": plan_label,
            "subscription_start_date": subscription.started_at.isoformat(),
            "subscription_end_date": subscription.expires_at.isoformat(),
            "days_remaining": days_remaining,
            "is_custom_trial": subscription.plan == SubscriptionPlan.STARTER and not subscription.billing_key,
            "billing_amount": int(_safe_decimal_to_int(next((p.amount for p in payments if p.status == PaymentStatus.PAID), 0))),
            "seats_used": 0,
            "seats_total": 999,
            "scheduled_plan": None,
            "scheduled_plan_date": None,
            "history": [
                {
                    "id": str(p.id),
                    "date": p.created_at.strftime("%Y-%m-%d"),
                    "description": f"{plan_label} 구독 결제",
                    "amount": int(_safe_decimal_to_int(p.amount)),
                    "status": "paid" if p.status == PaymentStatus.PAID else "failed",
                }
                for p in payments
            ],
        }
    )


@router.post("/billing/payment-method", response_model=APIResponse[dict])
async def change_payment_method(
    payload: PaymentMethodRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    org_id = _require_org_id(current_user)

    subscription = (await db.execute(select(Subscription).where(Subscription.organization_id == org_id))).scalar_one_or_none()
    if not subscription:
        raise HTTPException(status_code=404, detail="구독 정보를 찾을 수 없어요")

    subscription.billing_key = f"manual_{payload.card_number[-4:]}_{payload.expiry}"
    subscription.updated_at = datetime.utcnow()

    return APIResponse.ok(
        {
            "billing_key": subscription.billing_key,
            "payment_method": {
                "brand": "card",
                "last4": payload.card_number[-4:],
                "expires": payload.expiry,
            },
            "message": "결제수단을 변경했어요.",
        }
    )


@router.get("/admin/sa-dashboard", response_model=APIResponse[dict])
async def get_sa_dashboard(
    db: DBSession,
    current_user: CurrentUser,
):
    if _role_value(current_user) != "super_admin":
        raise HTTPException(status_code=403, detail="슈퍼관리자만 접근 가능해요")

    total_tenants = (await db.execute(select(func.count()).select_from(Organization))).scalar() or 0
    total_users = (await db.execute(select(func.count()).select_from(User))).scalar() or 0

    monthly_revenue = (await db.execute(
        select(func.coalesce(func.sum(Payment.amount), 0)).where(Payment.status == PaymentStatus.PAID)
    )).scalar() or Decimal("0")

    recent_orgs = (await db.execute(
        select(Organization).order_by(Organization.created_at.desc()).limit(5)
    )).scalars().all()

    plan_counts_raw = (await db.execute(
        select(Subscription.plan, func.count())
        .group_by(Subscription.plan)
    )).all()
    plan_distribution = []
    for plan, count in plan_counts_raw:
        plan_distribution.append({
            "plan": plan.value,
            "count": count,
            "percentage": round((count / total_tenants * 100), 1) if total_tenants else 0,
        })

    # UI currently expects nested dashboard payload.
    stats_payload = {
        "total_tenants": total_tenants,
        "total_users": total_users,
        "monthly_revenue": int(monthly_revenue),
        "new_signups": len(recent_orgs),
        "tenants_growth": 0,
        "users_growth": 0,
        "revenue_growth": 0,
        "signups_growth": 0,
    }

    # Build last 6 months of revenue history
    now_dt = datetime.utcnow()
    monthly_revenue_list = []
    for i in range(5, -1, -1):
        m_date = now_dt - timedelta(days=30 * i)
        monthly_revenue_list.append(
            {
                "month": m_date.strftime("%Y-%m"),
                "label": m_date.strftime("%m월"),
                "amount": int(monthly_revenue) if i == 0 else 0,
            }
        )

    return APIResponse.ok(
        {
            # Flattened keys for API client compatibility
            "total_tenants": stats_payload["total_tenants"],
            "active_tenants": stats_payload["total_tenants"],
            "total_revenue": stats_payload["monthly_revenue"],
            "recent_signups": [
                {
                    "id": str(org.id),
                    "company_name": org.name,
                    "plan": "starter",
                    "created_at": org.created_at.isoformat(),
                }
                for org in recent_orgs
            ],
            "subscription_breakdown": {p["plan"]: p["count"] for p in plan_distribution},
            # Nested keys for current dashboard page compatibility
            "stats": stats_payload,
            "recent_activity": [
                {
                    "id": str(org.id),
                    "type": "tenant",
                    "title": "신규 고객사 등록",
                    "description": org.name,
                    "timestamp": org.created_at.isoformat(),
                }
                for org in recent_orgs
            ],
            "monthly_revenue": monthly_revenue_list,
            "plan_distribution": plan_distribution,
        }
    )


@router.get("/admin/subscriptions/expiring", response_model=APIResponse[dict])
async def get_expiring_subscriptions(
    db: DBSession,
    current_user: CurrentUser,
    days: int = 30,
):
    """만료 임박 구독 목록 (super admin 전용)."""
    if _role_value(current_user) != "super_admin":
        raise HTTPException(status_code=403, detail="슈퍼관리자만 접근 가능해요")

    cutoff = datetime.utcnow() + timedelta(days=days)
    expiring_subs = (
        await db.execute(
            select(Subscription, Organization)
            .join(Organization, Subscription.organization_id == Organization.id)
            .where(
                Subscription.expires_at <= cutoff,
                Subscription.status == SubscriptionStatus.ACTIVE,
            )
            .order_by(Subscription.expires_at.asc())
        )
    ).all()

    items = []
    for sub, org in expiring_subs:
        days_remaining = max(0, (sub.expires_at.date() - datetime.utcnow().date()).days)
        plan_map = {
            SubscriptionPlan.STARTER: "스타터",
            SubscriptionPlan.STANDARD: "스탠다드",
            SubscriptionPlan.PREMIUM: "프리미엄",
        }
        items.append(
            {
                "id": str(sub.id),
                "company_name": org.name,
                "plan": plan_map.get(sub.plan, sub.plan.value),
                "expires_at": sub.expires_at.strftime("%Y-%m-%d"),
                "days_remaining": days_remaining,
            }
        )

    return APIResponse.ok({"items": items, "total": len(items)})


@router.get("/admin/tenants", response_model=PaginatedResponse[dict])
async def list_tenants(
    db: DBSession,
    current_user: CurrentUser,
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=20, ge=1, le=100),
    search: Optional[str] = None,
):
    if _role_value(current_user) != "super_admin":
        raise HTTPException(status_code=403, detail="슈퍼관리자만 접근 가능해요")

    query = select(Organization)
    if search:
        query = query.where(Organization.name.ilike(f"%{search}%"))

    total = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar() or 0
    orgs = (await db.execute(
        query.order_by(Organization.created_at.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
    )).scalars().all()

    items = []
    for org in orgs:
        users_count = (await db.execute(select(func.count()).select_from(User).where(User.organization_id == org.id))).scalar() or 0
        projects_count = (await db.execute(select(func.count()).select_from(Project).where(Project.organization_id == org.id))).scalar() or 0
        subscription = (await db.execute(select(Subscription).where(Subscription.organization_id == org.id))).scalar_one_or_none()

        plan = "trial"
        if subscription:
            if subscription.plan == SubscriptionPlan.STARTER:
                plan = "basic"
            elif subscription.plan == SubscriptionPlan.STANDARD:
                plan = "pro"
            elif subscription.plan == SubscriptionPlan.PREMIUM:
                plan = "pro"

        items.append(
            {
                "id": str(org.id),
                "name": org.name,
                "plan": plan,
                "users_count": users_count,
                "projects_count": projects_count,
                "created_at": org.created_at.isoformat(),
                "billing_amount": 0,
            }
        )

    return PaginatedResponse.create(items=items, page=page, per_page=per_page, total=total)


@router.get("/admin/tenants/{tenant_id}", response_model=APIResponse[dict])
async def get_tenant(
    tenant_id: int,
    db: DBSession,
    current_user: CurrentUser,
):
    if _role_value(current_user) != "super_admin":
        raise HTTPException(status_code=403, detail="슈퍼관리자만 접근 가능해요")

    org = (await db.execute(select(Organization).where(Organization.id == tenant_id))).scalar_one_or_none()
    if not org:
        raise NotFoundException("tenant", tenant_id)

    users_count = (await db.execute(select(func.count()).select_from(User).where(User.organization_id == org.id))).scalar() or 0
    projects_count = (await db.execute(select(func.count()).select_from(Project).where(Project.organization_id == org.id))).scalar() or 0

    subscription = (await db.execute(select(Subscription).where(Subscription.organization_id == org.id))).scalar_one_or_none()

    now = datetime.utcnow()
    if subscription:
        plan = "basic" if subscription.plan == SubscriptionPlan.STARTER else "pro"
        start_date = subscription.started_at.isoformat()
        end_date = subscription.expires_at.isoformat()
        billing_amount = int(
            _safe_decimal_to_int(
                (await db.execute(
                    select(func.coalesce(func.sum(Payment.amount), 0)).where(Payment.organization_id == org.id)
                )).scalar() or 0
            )
        )
    else:
        plan = "trial"
        start_date = org.created_at.isoformat()
        end_date = (org.created_at + timedelta(days=30)).isoformat()
        billing_amount = 0

    return APIResponse.ok(
        {
            "id": str(org.id),
            "name": org.name,
            "plan": plan,
            "users_count": users_count,
            "projects_count": projects_count,
            "created_at": org.created_at.isoformat(),
            "business_number": org.business_number,
            "representative": org.rep_name,
            "rep_phone": org.rep_phone,
            "rep_email": org.rep_email,
            "contact_name": org.contact_name,
            "contact_phone": org.contact_phone,
            "contact_position": org.contact_position,
            "subscription_start_date": start_date,
            "subscription_end_date": end_date,
            "is_custom_trial": plan == "trial",
            "billing_amount": billing_amount,
            "is_active": org.is_active,
        }
    )


@router.post("/admin/tenants/{tenant_id}/custom-trial", response_model=APIResponse[dict])
async def set_custom_trial_period(
    tenant_id: int,
    payload: CustomTrialRequest,
    db: DBSession,
    current_user: CurrentUser,
):
    if _role_value(current_user) != "super_admin":
        raise HTTPException(status_code=403, detail="슈퍼관리자만 접근 가능해요")

    org = (await db.execute(select(Organization).where(Organization.id == tenant_id))).scalar_one_or_none()
    if not org:
        raise NotFoundException("tenant", tenant_id)

    end_date = datetime.fromisoformat(payload.end_date.replace("Z", "+00:00")).replace(tzinfo=None)

    sub = (await db.execute(select(Subscription).where(Subscription.organization_id == org.id))).scalar_one_or_none()
    if not sub:
        sub = Subscription(
            organization_id=org.id,
            plan=SubscriptionPlan.STARTER,
            status=SubscriptionStatus.ACTIVE,
            started_at=datetime.utcnow(),
            expires_at=end_date,
        )
        db.add(sub)
        await db.flush()
    else:
        sub.expires_at = end_date
        sub.updated_at = datetime.utcnow()

    return APIResponse.ok(
        {
            "id": str(org.id),
            "subscription_end_date": end_date.isoformat(),
            "is_custom_trial": True,
        }
    )
