"""건설 원가계산서 엑셀 내보내기 서비스.

Generates Excel workbooks matching the official 원가계산서 format.
Uses openpyxl for Excel generation.
"""
from decimal import Decimal
from typing import Optional
import io

try:
    import openpyxl
    from openpyxl.styles import (
        Font, Alignment, Border, Side, PatternFill, numbers
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def _fmt_num(value) -> str:
    """Format number with Korean comma notation."""
    if value is None:
        return "0"
    try:
        return f"{int(value):,}"
    except (TypeError, ValueError):
        return str(value)


def _thin_border():
    """Create thin border for cells."""
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center_align():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _right_align():
    return Alignment(horizontal="right", vertical="center")


def _left_align():
    return Alignment(horizontal="left", vertical="center")


def _header_fill():
    return PatternFill("solid", fgColor="D9E1F2")  # 연한 파란색


def _subtotal_fill():
    return PatternFill("solid", fgColor="E2EFDA")  # 연한 초록색


def _total_fill():
    return PatternFill("solid", fgColor="FFF2CC")  # 연한 노란색


def generate_cost_calculation_sheet(calc_data: dict, rate_data: dict) -> bytes:
    """
    원가계산서 엑셀 시트 생성.

    Args:
        calc_data: CostCalculation dict with all computed fields
        rate_data: ConstructionCostRate dict with rate values

    Returns:
        Excel file bytes
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl is not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "원가계산서"

    # Column widths
    ws.column_dimensions["A"].width = 6   # No
    ws.column_dimensions["B"].width = 30  # 항목명
    ws.column_dimensions["C"].width = 25  # 산출근거
    ws.column_dimensions["D"].width = 10  # 요율(%)
    ws.column_dimensions["E"].width = 18  # 금액

    border = _thin_border()
    header_fill = _header_fill()
    subtotal_fill = _subtotal_fill()
    total_fill = _total_fill()

    row = 1

    def add_header_row(label: str):
        nonlocal row
        ws.merge_cells(f"A{row}:E{row}")
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(bold=True, size=14)
        cell.alignment = _center_align()
        row += 1

    def add_section_header():
        nonlocal row
        headers = ["No", "항목명", "산출근거", "요율(%)", "금액(원)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = _center_align()
            cell.border = border
        row += 1

    def add_item_row(no: str, name: str, basis: str, rate: str, amount, is_subtotal=False, is_total=False, is_readonly=False):
        nonlocal row
        fill = None
        if is_total:
            fill = total_fill
        elif is_subtotal:
            fill = subtotal_fill

        values = [no, name, basis, rate, _fmt_num(amount)]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = _center_align() if col in (1, 3, 4) else _left_align() if col == 2 else _right_align()
            if fill:
                cell.fill = fill
            if is_subtotal or is_total:
                cell.font = Font(bold=True)
        row += 1

    # Title
    add_header_row("원   가   계   산   서")
    row += 1

    # Column headers
    add_section_header()

    # --- 재료비 ---
    add_item_row("", "【재료비】", "", "", "", is_subtotal=True)
    add_item_row("1", "직접재료비", "", "", calc_data.get("direct_material_cost", 0))
    add_item_row("2", "간접재료비", "", "", calc_data.get("indirect_material_cost", 0))
    scrap = calc_data.get("material_scrap", 0)
    if scrap and int(scrap) != 0:
        add_item_row("3", "작업설 및 부산물 (△)", "", "", f"△{_fmt_num(scrap)}")
    add_item_row("", "재료비 소계", "", "", calc_data.get("material_subtotal", 0), is_subtotal=True)

    row += 1  # spacing

    # --- 노무비 ---
    add_item_row("", "【노무비】", "", "", "", is_subtotal=True)
    add_item_row("", "직접노무비", "", "", calc_data.get("direct_labor_cost", 0))
    indirect_rate = calc_data.get("override_indirect_labor_rate") or rate_data.get("indirect_labor_rate", "15.00")
    add_item_row("", "간접노무비", f"직접노무비 × {indirect_rate}%", f"{indirect_rate}%", calc_data.get("indirect_labor_amount", 0))
    add_item_row("", "노무비 소계", "", "", calc_data.get("labor_subtotal", 0), is_subtotal=True)

    row += 1

    # --- 경비 ---
    add_item_row("", "【경비】", "", "", "", is_subtotal=True)
    add_item_row("", "기계경비", "", "", calc_data.get("equipment_cost", 0))
    add_item_row("", "산재보험료", f"노무비소계 × {rate_data.get('industrial_accident_rate', '3.56')}%", f"{rate_data.get('industrial_accident_rate', '3.56')}%", calc_data.get("accident_insurance", 0), is_readonly=True)
    add_item_row("", "고용보험료", f"노무비소계 × {rate_data.get('employment_insurance_rate', '1.01')}%", f"{rate_data.get('employment_insurance_rate', '1.01')}%", calc_data.get("employment_insurance", 0), is_readonly=True)

    days = calc_data.get("construction_days", 30)
    min_days = rate_data.get("health_insurance_min_days", 31)
    health_note = f"직접노무비 × {rate_data.get('health_insurance_rate', '3.545')}%" if days >= min_days else f"공사기간 {days}일 < {min_days}일 미적용"
    add_item_row("", "건강보험료", health_note, f"{rate_data.get('health_insurance_rate', '3.545')}%" if days >= min_days else "-", calc_data.get("health_insurance", 0), is_readonly=True)
    add_item_row("", "국민연금보험료", "", f"{rate_data.get('national_pension_rate', '4.50')}%", calc_data.get("national_pension", 0), is_readonly=True)
    add_item_row("", "노인장기요양보험료", f"건강보험료 × {rate_data.get('longterm_care_rate', '12.95')}%", f"{rate_data.get('longterm_care_rate', '12.95')}%", calc_data.get("longterm_care", 0), is_readonly=True)
    add_item_row("", "산업안전보건관리비", f"(재료비+직접노무비) × {rate_data.get('safety_management_rate', '3.11')}%", f"{rate_data.get('safety_management_rate', '3.11')}%", calc_data.get("safety_management", 0), is_readonly=True)
    add_item_row("", "환경보전비", f"(재료비+노무비+기계경비) × {rate_data.get('environmental_rate', '0.30')}%", f"{rate_data.get('environmental_rate', '0.30')}%", calc_data.get("environmental_fee", 0), is_readonly=True)
    other_rate = calc_data.get("override_other_expense_rate") or rate_data.get("other_expense_rate", "4.60")
    add_item_row("", "기타경비", f"(재료비+노무비) × {other_rate}%", f"{other_rate}%", calc_data.get("other_expense", 0))

    if calc_data.get("enable_subcontract_guarantee"):
        add_item_row("", "하도급보증수수료", f"{rate_data.get('subcontract_guarantee_rate', '0.081')}%", f"{rate_data.get('subcontract_guarantee_rate', '0.081')}%", calc_data.get("subcontract_guarantee", 0))
    if calc_data.get("enable_equipment_guarantee"):
        add_item_row("", "건설기계보증수수료", f"직접노무비 × {rate_data.get('equipment_guarantee_rate', '0.10')}%", f"{rate_data.get('equipment_guarantee_rate', '0.10')}%", calc_data.get("equipment_guarantee", 0))

    add_item_row("", "경비 소계", "", "", calc_data.get("expense_subtotal", 0), is_subtotal=True)

    row += 1

    # --- 순공사원가 ---
    add_item_row("", "순공사원가", "재료비 + 노무비 + 경비", "", calc_data.get("net_construction_cost", 0), is_subtotal=True)

    row += 1

    # --- 일반관리비 & 이윤 ---
    admin_rate = calc_data.get("override_general_admin_rate") or rate_data.get("general_admin_rate", "5.50")
    add_item_row("", "일반관리비", f"순공사원가 × {admin_rate}%", f"{admin_rate}%", calc_data.get("general_admin_fee", 0))

    waste = calc_data.get("waste_disposal_fee", 0)
    if waste and int(waste) != 0:
        add_item_row("", "폐기물처리비", "", "", waste)

    profit_rate = calc_data.get("override_profit_rate") or rate_data.get("profit_rate_cap", "12.00")
    adj = calc_data.get("profit_adjustment", 0)
    profit_note = f"(노무비+경비+일반관리비) × {profit_rate}%"
    if adj and int(adj) != 0:
        profit_note += f" + 조정({_fmt_num(adj)})"
    add_item_row("", "이윤", profit_note, f"{profit_rate}%", calc_data.get("profit_amount", 0))

    row += 1

    # --- 최종 합계 ---
    add_item_row("", "공급가액", "", "", calc_data.get("supply_amount", 0), is_total=True)
    add_item_row("", "부가가치세 (10%)", "공급가액 × 10%", "10%", calc_data.get("vat_amount", 0), is_total=True)
    add_item_row("", "도급액 합계", "공급가액 + 부가가치세", "", calc_data.get("contract_amount", 0), is_total=True)

    # Freeze header rows
    ws.freeze_panes = f"A3"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_combined_export(estimate_data: dict, lines_data: list, calc_data: dict, rate_data: dict) -> bytes:
    """
    견적서 + 원가계산서 합본 엑셀 생성.

    Args:
        estimate_data: Estimate dict
        lines_data: List of EstimateLine dicts
        calc_data: CostCalculation dict
        rate_data: ConstructionCostRate dict

    Returns:
        Excel file bytes with 2 sheets: 내역서, 원가계산서
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl is not installed")

    wb = openpyxl.Workbook()

    # Sheet 1: 내역서 (Estimate lines)
    ws1 = wb.active
    ws1.title = "내역서"

    border = _thin_border()
    header_fill = _header_fill()

    # 내역서 headers
    headers = ["No", "품목명", "규격", "단위", "수량", "단가(원)", "재료비", "노무비", "경비", "금액(원)"]
    col_widths = [6, 30, 20, 8, 10, 15, 15, 15, 15, 15]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws1.column_dimensions[get_column_letter(col)].width = w
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = _center_align()
        cell.border = border

    for i, line in enumerate(lines_data, 1):
        row_data = [
            i,
            line.get("description", ""),
            line.get("specification", ""),
            line.get("unit", ""),
            _fmt_num(line.get("quantity", 0)),
            _fmt_num(line.get("unit_price_snapshot", 0)),
            _fmt_num(line.get("material_amount", 0)),
            _fmt_num(line.get("labor_amount", 0)),
            _fmt_num(line.get("equipment_amount", 0)),
            _fmt_num(line.get("amount", 0)),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=i+1, column=col, value=val)
            cell.border = border
            cell.alignment = _right_align() if col in (5, 6, 7, 8, 9, 10) else _left_align()

    # Totals row
    tot_row = len(lines_data) + 2
    ws1.cell(row=tot_row, column=2, value="소 계").font = Font(bold=True)
    ws1.cell(row=tot_row, column=10, value=_fmt_num(estimate_data.get("subtotal", 0))).font = Font(bold=True)

    # Sheet 2: 원가계산서
    ws2 = wb.create_sheet(title="원가계산서")

    # Generate cost calc content in ws2
    # Reuse the single-sheet logic but write to ws2
    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 25
    ws2.column_dimensions["D"].width = 10
    ws2.column_dimensions["E"].width = 18

    # Title
    ws2.merge_cells("A1:E1")
    title_cell = ws2.cell(row=1, column=1, value="원   가   계   산   서")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = _center_align()

    # Add key summary items only in combined export (detailed items in separate sheet)
    summary_items = [
        ("", "직접재료비", "", "", calc_data.get("direct_material_cost", 0)),
        ("", "직접노무비", "", "", calc_data.get("direct_labor_cost", 0)),
        ("", "기계경비", "", "", calc_data.get("equipment_cost", 0)),
        ("", "재료비 소계", "", "", calc_data.get("material_subtotal", 0)),
        ("", "노무비 소계", "", "", calc_data.get("labor_subtotal", 0)),
        ("", "경비 소계", "", "", calc_data.get("expense_subtotal", 0)),
        ("", "순공사원가", "", "", calc_data.get("net_construction_cost", 0)),
        ("", "일반관리비", "", "", calc_data.get("general_admin_fee", 0)),
        ("", "이윤", "", "", calc_data.get("profit_amount", 0)),
        ("", "공급가액", "", "", calc_data.get("supply_amount", 0)),
        ("", "부가가치세 (10%)", "", "", calc_data.get("vat_amount", 0)),
        ("", "도급액 합계", "", "", calc_data.get("contract_amount", 0)),
    ]

    for r, (no, name, basis, rate, amount) in enumerate(summary_items, 3):
        is_key = name in ("재료비 소계", "노무비 소계", "경비 소계", "순공사원가", "공급가액", "도급액 합계")
        fill = _subtotal_fill() if is_key else None
        is_total_row = name in ("공급가액", "부가가치세 (10%)", "도급액 합계")
        if is_total_row:
            fill = _total_fill()
        for col, val in enumerate([no, name, basis, rate, _fmt_num(amount)], 1):
            cell = ws2.cell(row=r, column=col, value=val)
            cell.border = border
            cell.alignment = _center_align() if col in (1, 3, 4) else _left_align() if col == 2 else _right_align()
            if fill:
                cell.fill = fill
            if is_key or is_total_row:
                cell.font = Font(bold=True)

    # Save
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
